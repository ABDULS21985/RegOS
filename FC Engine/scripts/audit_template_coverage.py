#!/usr/bin/env python3
"""Audit coverage between template source artifacts and module definition JSON files.

Checks completed modules:
- RG-08: BDC_CBN, MFB_PAR, NFIU_AML
- RG-09: DMB_BASEL3, NDIC_RETURNS, PSP_FINTECH, PMB_CBN
- RG-10: INSURANCE_NAICOM, PFA_PENCOM, CMO_SEC, DFI_CBN, IMTO_CBN
- RG-11: FATF_EVAL

Compares:
- Workbook operational sheets vs JSON template count
- Workbook Data Dictionary fields vs JSON field codes (normalized exact)
- XSD element names vs JSON field codes (normalized exact)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import xml.etree.ElementTree as et
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class ModuleSpec:
    module_code: str
    workbook: str
    xsd: str
    definition_json: str


MODULES: tuple[ModuleSpec, ...] = (
    ModuleSpec(
        module_code="BDC_CBN",
        workbook="Templates/BDC_Reporting_Templates.xlsx",
        xsd="Templates/BDC_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/src/FC.Engine.Migrator/SeedData/ModuleDefinitions/rg08-bdc-cbn-module-definition.json",
    ),
    ModuleSpec(
        module_code="MFB_PAR",
        workbook="Templates/MFB_Reporting_Templates.xlsx",
        xsd="Templates/MFB_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/src/FC.Engine.Migrator/SeedData/ModuleDefinitions/rg08-mfb-par-module-definition.json",
    ),
    ModuleSpec(
        module_code="NFIU_AML",
        workbook="Templates/NFIU_Reporting_Templates.xlsx",
        xsd="Templates/NFIU_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/src/FC.Engine.Migrator/SeedData/ModuleDefinitions/rg08-nfiu-aml-module-definition.json",
    ),
    ModuleSpec(
        module_code="DMB_BASEL3",
        workbook="Templates/DMB_Reporting_Templates.xlsx",
        xsd="Templates/DMB_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg09/dmb_basel3.json",
    ),
    ModuleSpec(
        module_code="NDIC_RETURNS",
        workbook="Templates/NDIC_Returns_Reporting_Templates.xlsx",
        xsd="Templates/NDIC_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg09/ndic_returns.json",
    ),
    ModuleSpec(
        module_code="PSP_FINTECH",
        workbook="Templates/PSP_Fintech_Reporting_Templates.xlsx",
        xsd="Templates/PSP_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg09/psp_fintech.json",
    ),
    ModuleSpec(
        module_code="PMB_CBN",
        workbook="Templates/PMB_CBN_FMBN_Reporting_Templates.xlsx",
        xsd="Templates/PMB_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg09/pmb_cbn.json",
    ),
    ModuleSpec(
        module_code="INSURANCE_NAICOM",
        workbook="Templates/Insurance_NAICOM_Reporting_Templates.xlsx",
        xsd="Templates/Insurance_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg10/insurance_naicom.json",
    ),
    ModuleSpec(
        module_code="PFA_PENCOM",
        workbook="Templates/PFA_PenCom_Reporting_Templates.xlsx",
        xsd="Templates/PFA_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg10/pfa_pencom.json",
    ),
    ModuleSpec(
        module_code="CMO_SEC",
        workbook="Templates/Capital_Market_SEC_Reporting_Templates.xlsx",
        xsd="Templates/SEC_CMO_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg10/cmo_sec.json",
    ),
    ModuleSpec(
        module_code="DFI_CBN",
        workbook="Templates/DFI_CBN_Reporting_Templates.xlsx",
        xsd="Templates/DFI_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg10/dfi_cbn.json",
    ),
    ModuleSpec(
        module_code="IMTO_CBN",
        workbook="Templates/IMTO_CBN_Reporting_Templates.xlsx",
        xsd="Templates/IMTO_Return_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg10/imto_cbn.json",
    ),
    ModuleSpec(
        module_code="FATF_EVAL",
        workbook="Templates/FATF_Mutual_Evaluation_Templates.xlsx",
        xsd="Templates/FATF_ME_Schema_v1.0.xsd",
        definition_json="FC Engine/docs/module-definitions/rg11/fatf_eval.json",
    ),
)


def norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def find_repo_root(start: Path) -> Path:
    cur = start.resolve()
    for candidate in (cur, *cur.parents):
        if (candidate / ".git").exists() and (candidate / "FC Engine").exists():
            return candidate
    raise FileNotFoundError("Could not locate repo root from current directory.")


def read_workbook_fields_and_sheets(path: Path) -> tuple[list[str], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    dd_sheet = "Data Dictionary" if "Data Dictionary" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[dd_sheet]

    fields: list[str] = []
    for row_idx in range(1, ws.max_row + 1):
        serial = ws.cell(row_idx, 1).value
        field = ws.cell(row_idx, 2).value
        if serial is None or field is None:
            continue
        if str(serial).strip().isdigit():
            fields.append(str(field).strip())

    operational_sheets = [
        s
        for s in wb.sheetnames
        if "cover" not in s.lower()
        and "dictionary" not in s.lower()
        and "schema" not in s.lower()
        and "reference" not in s.lower()
    ]

    wb.close()
    return fields, operational_sheets


def read_xsd_elements(path: Path) -> list[str]:
    tree = et.parse(path)
    root = tree.getroot()
    ns = "{http://www.w3.org/2001/XMLSchema}"

    names: list[str] = []
    for elem in root.iter(f"{ns}element"):
        name = elem.attrib.get("name")
        if not name:
            continue
        # Exclude wrapper names; keep business fields.
        if name.endswith("Return") or name.startswith("Field_"):
            continue
        names.append(name)

    # Preserve order but dedupe.
    seen: set[str] = set()
    unique: list[str] = []
    for n in names:
        if n in seen:
            continue
        seen.add(n)
        unique.append(n)
    return unique


def read_definition(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def exact_normalized_matches(source: Iterable[str], target: Iterable[str]) -> tuple[list[str], list[str]]:
    target_norm = {norm(t): t for t in target}
    matched: list[str] = []
    missing: list[str] = []
    for s in source:
        if norm(s) in target_norm:
            matched.append(s)
        else:
            missing.append(s)
    return matched, missing


def render_report(root: Path) -> str:
    lines: list[str] = []
    lines.append("# Template Coverage Audit")
    lines.append("")
    lines.append("Generated from `/Templates` workbook/XSD artifacts vs imported module-definition JSON files.")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append("| Module | WB Sheets | JSON Templates | WB Fields | JSON Fields | WB Field Match | XSD Elements | XSD Field Match |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|")

    detail_blocks: list[str] = []

    for spec in MODULES:
        wb_path = root / spec.workbook
        xsd_path = root / spec.xsd
        json_path = root / spec.definition_json

        if not wb_path.exists() or not xsd_path.exists() or not json_path.exists():
            detail_blocks.append(
                f"## {spec.module_code}\n\nMissing source artifact(s): "
                f"workbook={wb_path.exists()} xsd={xsd_path.exists()} json={json_path.exists()}\n"
            )
            continue

        wb_fields, wb_sheets = read_workbook_fields_and_sheets(wb_path)
        xsd_elements = read_xsd_elements(xsd_path)

        definition = read_definition(json_path)
        templates = definition.get("templates", [])
        json_fields = [f["fieldCode"] for t in templates for f in t.get("fields", [])]

        wb_matched, wb_missing = exact_normalized_matches(wb_fields, json_fields)
        xsd_matched, xsd_missing = exact_normalized_matches(xsd_elements, json_fields)

        lines.append(
            f"| {spec.module_code} | {len(wb_sheets)} | {len(templates)} | {len(wb_fields)} | {len(json_fields)} | "
            f"{len(wb_matched)}/{len(wb_fields)} | {len(xsd_elements)} | {len(xsd_matched)}/{len(xsd_elements)} |"
        )

        detail_lines: list[str] = []
        detail_lines.append(f"## {spec.module_code}")
        detail_lines.append("")
        detail_lines.append(f"- Workbook: `{spec.workbook}`")
        detail_lines.append(f"- XSD: `{spec.xsd}`")
        detail_lines.append(f"- Definition: `{spec.definition_json}`")
        detail_lines.append(f"- JSON formulas: {sum(len(t.get('formulas', [])) for t in templates)}")
        detail_lines.append(f"- JSON cross-sheet rules: {sum(len(t.get('crossSheetRules', [])) for t in templates)}")
        detail_lines.append(f"- JSON inter-module flows: {len(definition.get('interModuleDataFlows', []))}")
        detail_lines.append("")
        detail_lines.append("Top workbook fields not matched (normalized exact):")
        detail_lines.append("")
        for item in wb_missing[:20]:
            detail_lines.append(f"- {item}")
        if not wb_missing:
            detail_lines.append("- None")
        detail_lines.append("")
        detail_lines.append("Top XSD elements not matched (normalized exact):")
        detail_lines.append("")
        for item in xsd_missing[:20]:
            detail_lines.append(f"- {item}")
        if not xsd_missing:
            detail_lines.append("- None")
        detail_lines.append("")
        detail_blocks.append("\n".join(detail_lines))

    lines.append("")
    lines.append("## Details")
    lines.append("")
    lines.extend(detail_blocks)

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit template coverage against module definition JSON files.")
    parser.add_argument(
        "--output",
        default="FC Engine/docs/template-coverage-audit.md",
        help="Output markdown report path relative to repo root.",
    )
    args = parser.parse_args()

    repo_root = find_repo_root(Path.cwd())
    report = render_report(repo_root)

    output_path = repo_root / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report, encoding="utf-8")

    print(f"Wrote report: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
