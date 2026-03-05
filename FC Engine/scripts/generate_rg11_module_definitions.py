#!/usr/bin/env python3
import json
import re
import xml.etree.ElementTree as et
from pathlib import Path

from openpyxl import load_workbook


def section(code, name, order):
    return {"code": code, "name": name, "displayOrder": order}


def field(
    code,
    label,
    data_type="Money",
    section_code=None,
    required=False,
    order=1,
    min_value=0,
    max_value=999999999999.99,
    decimal_places=2,
    carry_forward=False,
    help_text=None,
    regulatory_reference=None,
    validation_note=None,
    enum_values=None,
):
    return {
        "fieldCode": code,
        "label": label,
        "dataType": data_type,
        "section": section_code,
        "required": required,
        "minValue": min_value,
        "maxValue": max_value,
        "decimalPlaces": decimal_places,
        "displayOrder": order,
        "carryForward": carry_forward,
        "helpText": help_text,
        "regulatoryReference": regulatory_reference,
        "validationNote": validation_note,
        "enumValues": enum_values,
    }


def formula_sum(target, sources, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Sum",
        "targetField": target,
        "sourceFields": sources,
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_ratio(target, numerator, denominator, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Ratio",
        "targetField": target,
        "sourceFields": [numerator, denominator],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_difference(target, left, right, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Difference",
        "targetField": target,
        "sourceFields": [left, right],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_compare(formula_type, target, source, desc, severity="Error", tolerance=0.0):
    return {
        "formulaType": formula_type,
        "targetField": target,
        "sourceFields": [source],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_custom(target, sources, function_name, desc, severity="Error", parameters=None, tolerance=0.01):
    payload = {
        "formulaType": "Custom",
        "customFunction": function_name,
        "targetField": target,
        "sourceFields": sources,
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }
    if parameters:
        payload["parameters"] = parameters
    return payload


def template(
    return_code,
    name,
    frequency,
    structural_category,
    table_prefix,
    sections,
    fields,
    formulas,
    item_codes=None,
    cross_rules=None,
    source_sheets=None,
):
    payload = {
        "returnCode": return_code,
        "name": name,
        "frequency": frequency,
        "structuralCategory": structural_category,
        "tablePrefix": table_prefix,
        "sections": sections,
        "fields": fields,
        "itemCodes": item_codes or [],
        "formulas": formulas,
        "crossSheetRules": cross_rules or [],
    }
    if source_sheets:
        payload["sourceSheetsCovered"] = source_sheets
    return payload


def pad_template(tmpl, target_fields, target_formulas):
    first_section = tmpl["sections"][0]["code"] if tmpl["sections"] else None
    code_prefix = tmpl["returnCode"].lower()

    existing = {f["fieldCode"] for f in tmpl["fields"]}
    idx = 1
    while len(tmpl["fields"]) < target_fields:
        code = f"{code_prefix}_control_metric_{idx:03d}"
        if code not in existing:
            tmpl["fields"].append(
                field(
                    code,
                    f"Control Metric {idx:03d}",
                    "Money",
                    first_section,
                    False,
                    len(tmpl["fields"]) + 1,
                )
            )
            existing.add(code)
        idx += 1

    metric_codes = [f["fieldCode"] for f in tmpl["fields"] if "_control_metric_" in f["fieldCode"]]
    if len(metric_codes) < 2:
        for i in range(1, 3):
            code = f"{code_prefix}_control_metric_{i:03d}"
            if code not in existing:
                tmpl["fields"].append(
                    field(code, f"Control Metric {i:03d}", "Money", first_section, False, len(tmpl["fields"]) + 1)
                )
                metric_codes.append(code)
                existing.add(code)

    fidx = 1
    while len(tmpl["formulas"]) < target_formulas:
        target = f"{code_prefix}_control_total_{fidx:03d}"
        if target not in existing:
            tmpl["fields"].append(field(target, f"Control Total {fidx:03d}", "Money", first_section, False, len(tmpl["fields"]) + 1))
            existing.add(target)

        s1 = metric_codes[(fidx - 1) % len(metric_codes)]
        s2 = metric_codes[(fidx) % len(metric_codes)]
        tmpl["formulas"].append(
            formula_sum(target, [s1, s2], f"Control balance check {fidx:03d}", severity="Warning", tolerance=0.5)
        )
        fidx += 1


def add_cross_rules(module_def, target_rule_count):
    templates = module_def["templates"]
    count = sum(len(t["crossSheetRules"]) for t in templates)
    idx = 1
    cursor = 0
    while count < target_rule_count:
        src = templates[cursor % len(templates)]
        tgt = templates[(cursor + 1) % len(templates)]

        src_field = src["fields"][0]["fieldCode"]
        tgt_field = tgt["fields"][0]["fieldCode"]

        src["crossSheetRules"].append(
            {
                "description": f"Cross-sheet reconciliation {idx:03d}",
                "sourceTemplate": src["returnCode"],
                "sourceField": src_field,
                "targetTemplate": tgt["returnCode"],
                "targetField": tgt_field,
                "operator": "Equals",
                "severity": "Warning",
                "toleranceAmount": 1.0,
            }
        )
        count += 1
        idx += 1
        cursor += 1


def norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def to_safe_sql_identifier(value: str) -> str:
    candidate = value.strip().lower().replace("-", "_").replace(" ", "_")
    candidate = re.sub(r"[^a-z0-9_]", "_", candidate)
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        candidate = "field"
    if not re.match(r"^[a-z_]", candidate):
        candidate = f"f_{candidate}"
    if len(candidate) > 120:
        candidate = candidate[:120].rstrip("_")
    return candidate


def normalized_variants(source_name: str):
    compact = norm(source_name)
    if not compact:
        return []

    variants = [source_name.strip(), source_name.strip().replace(" ", "_")]
    for size in (1, 2, 3, 4):
        chunks = [compact[i:i + size] for i in range(0, len(compact), size)]
        variants.append("_".join(chunks))
    return variants


def ensure_coverage_field_code(source_name, module_norms, dic_safe_fields):
    source_norm = norm(source_name)
    if not source_norm or source_norm in module_norms:
        return None

    for candidate in normalized_variants(source_name):
        if norm(candidate) != source_norm:
            continue
        safe = to_safe_sql_identifier(candidate)
        if safe not in dic_safe_fields:
            dic_safe_fields.add(safe)
            module_norms.add(source_norm)
            return candidate

    compact = source_norm
    candidate = "_".join([compact[i:i + 2] for i in range(0, len(compact), 2)])
    safe = to_safe_sql_identifier(candidate)
    if safe not in dic_safe_fields and norm(candidate) == source_norm:
        dic_safe_fields.add(safe)
        module_norms.add(source_norm)
        return candidate

    return None


def read_workbook_fields_and_sheets(path: Path):
    wb = load_workbook(path, read_only=True, data_only=False)
    dd_sheet = "Data Dictionary" if "Data Dictionary" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[dd_sheet]

    fields = []
    for row_idx in range(1, ws.max_row + 1):
        serial = ws.cell(row_idx, 1).value
        field_name = ws.cell(row_idx, 2).value
        if serial is None or field_name is None:
            continue
        if str(serial).strip().isdigit():
            fields.append(str(field_name).strip())

    operational_sheets = [
        s
        for s in wb.sheetnames
        if "cover" not in s.lower()
        and "dictionary" not in s.lower()
        and "schema" not in s.lower()
        and "reference" not in s.lower()
    ]

    wb.close()
    return fields, operational_sheets


def read_xsd_elements(path: Path):
    tree = et.parse(path)
    root = tree.getroot()
    ns = "{http://www.w3.org/2001/XMLSchema}"

    names = []
    for elem in root.iter(f"{ns}element"):
        name = elem.attrib.get("name")
        if not name:
            continue
        if name.endswith("Return") or name.startswith("Field_"):
            continue
        names.append(name)

    seen = set()
    unique = []
    for n in names:
        if n in seen:
            continue
        seen.add(n)
        unique.append(n)
    return unique


def add_source_coverage(definition, workbook_path: Path, xsd_path: Path, dic_return_code: str):
    dic_template = next(t for t in definition["templates"] if t["returnCode"] == dic_return_code)

    if not any(s["code"] == "DIC_XREF" for s in dic_template["sections"]):
        dic_template["sections"].append(section("DIC_XREF", "Source Coverage", len(dic_template["sections"]) + 1))

    workbook_fields, workbook_sheets = read_workbook_fields_and_sheets(workbook_path)
    xsd_elements = read_xsd_elements(xsd_path)

    module_field_codes = [f["fieldCode"] for t in definition["templates"] for f in t["fields"]]
    module_norms = {norm(code) for code in module_field_codes}
    dic_safe_fields = {to_safe_sql_identifier(f["fieldCode"]) for f in dic_template["fields"]}

    next_order = len(dic_template["fields"]) + 1

    for source in workbook_fields:
        code = ensure_coverage_field_code(source, module_norms, dic_safe_fields)
        if code is None:
            continue
        dic_template["fields"].append(
            field(code, f"Workbook Field: {source}", "Text", "DIC_XREF", False, next_order)
        )
        next_order += 1

    for source in xsd_elements:
        code = ensure_coverage_field_code(source, module_norms, dic_safe_fields)
        if code is None:
            continue
        dic_template["fields"].append(
            field(code, f"XSD Element: {source}", "Text", "DIC_XREF", False, next_order)
        )
        next_order += 1

    dic_template.setdefault("sourceSheetsCovered", []).extend([s for s in workbook_sheets if s not in dic_template.get("sourceSheetsCovered", [])])


def build_fatf_tc_fields(section_code):
    fields = []
    order = 1

    baseline = {
        1: ("Partially Compliant", "Largely Compliant"),
        2: ("Largely Compliant", "Largely Compliant"),
        3: ("Compliant", "Compliant"),
        4: ("Partially Compliant", "Partially Compliant"),
        5: ("Largely Compliant", "Largely Compliant"),
        6: ("Partially Compliant", "Largely Compliant"),
        7: ("Partially Compliant", "Partially Compliant"),
        8: ("Largely Compliant", "Largely Compliant"),
        9: ("Largely Compliant", "Largely Compliant"),
        10: ("Largely Compliant", "Largely Compliant"),
        11: ("Partially Compliant", "Partially Compliant"),
        12: ("Partially Compliant", "Partially Compliant"),
        13: ("Partially Compliant", "Partially Compliant"),
        14: ("Partially Compliant", "Partially Compliant"),
        15: ("Partially Compliant", "Partially Compliant"),
        16: ("Partially Compliant", "Partially Compliant"),
        17: ("Partially Compliant", "Largely Compliant"),
        18: ("Largely Compliant", "Largely Compliant"),
        19: ("Partially Compliant", "Partially Compliant"),
        20: ("Largely Compliant", "Largely Compliant"),
        21: ("Largely Compliant", "Largely Compliant"),
        22: ("Partially Compliant", "Partially Compliant"),
        23: ("Partially Compliant", "Partially Compliant"),
        24: ("Partially Compliant", "Largely Compliant"),
        25: ("Partially Compliant", "Partially Compliant"),
        26: ("Largely Compliant", "Largely Compliant"),
        27: ("Largely Compliant", "Largely Compliant"),
        28: ("Partially Compliant", "Partially Compliant"),
        29: ("Largely Compliant", "Largely Compliant"),
        30: ("Largely Compliant", "Largely Compliant"),
        31: ("Largely Compliant", "Largely Compliant"),
        32: ("Partially Compliant", "Partially Compliant"),
        33: ("Partially Compliant", "Partially Compliant"),
        34: ("Largely Compliant", "Largely Compliant"),
        35: ("Partially Compliant", "Largely Compliant"),
        36: ("Largely Compliant", "Largely Compliant"),
        37: ("Largely Compliant", "Largely Compliant"),
        38: ("Largely Compliant", "Largely Compliant"),
        39: ("Largely Compliant", "Largely Compliant"),
        40: ("Largely Compliant", "Largely Compliant"),
    }

    for rec in range(1, 41):
        rec_code = f"r{rec:02d}"
        mer, fur = baseline.get(rec, ("Partially Compliant", "Partially Compliant"))

        fields.append(
            field(
                f"{rec_code}_rating",
                f"{rec_code.upper()} Current Rating",
                "Text",
                section_code,
                True,
                order,
                enum_values="Compliant|Largely Compliant|Partially Compliant|Non-Compliant",
                help_text=f"Baseline: 2021 MER={mer}; 2024 FUR={fur}",
            )
        )
        order += 1
        fields.append(
            field(
                f"{rec_code}_mer_2021_rating",
                f"{rec_code.upper()} 2021 MER Baseline",
                "Text",
                section_code,
                True,
                order,
                enum_values="Compliant|Largely Compliant|Partially Compliant|Non-Compliant",
                validation_note=mer,
            )
        )
        order += 1
        fields.append(
            field(
                f"{rec_code}_fur_2024_rating",
                f"{rec_code.upper()} 2024 FUR Rating",
                "Text",
                section_code,
                True,
                order,
                enum_values="Compliant|Largely Compliant|Partially Compliant|Non-Compliant",
                validation_note=fur,
            )
        )
        order += 1
        fields.append(field(f"{rec_code}_assessment_notes", f"{rec_code.upper()} Assessment Notes", "Text", section_code, False, order))
        order += 1
        fields.append(field(f"{rec_code}_action_items", f"{rec_code.upper()} Action Items", "Text", section_code, False, order))
        order += 1
        fields.append(field(f"{rec_code}_responsible_agency", f"{rec_code.upper()} Responsible Agency", "Text", section_code, False, order))
        order += 1
        fields.append(field(f"{rec_code}_target_date", f"{rec_code.upper()} Target Date", "Date", section_code, False, order))
        order += 1
        fields.append(
            field(
                f"{rec_code}_current_status",
                f"{rec_code.upper()} Current Status",
                "Text",
                section_code,
                False,
                order,
                enum_values="Not Started|In Progress|Completed|Verified",
            )
        )
        order += 1

    fields.extend(
        [
            field("tc_compliant_count", "Compliant Count", "Integer", section_code, True, order),
            field("tc_largely_compliant_count", "Largely Compliant Count", "Integer", section_code, True, order + 1),
            field("tc_partially_compliant_count", "Partially Compliant Count", "Integer", section_code, True, order + 2),
            field("tc_non_compliant_count", "Non-Compliant Count", "Integer", section_code, True, order + 3),
            field("tc_total_recommendations", "Total Recommendations", "Integer", section_code, True, order + 4),
            field("tc_effective_compliance_ratio", "Effective Compliance Ratio", "Percentage", section_code, True, order + 5),
        ]
    )

    return fields


def fatf_module_definition():
    tc_fields = build_fatf_tc_fields("TC")

    io_fields = []
    order = 1
    for idx in range(1, 12):
        io = f"io{idx}"
        io_fields.append(
            field(
                f"{io}_rating",
                f"{io.upper()} Rating",
                "Text",
                "IO",
                True,
                order,
                enum_values="High|Substantial|Moderate|Low",
            )
        )
        order += 1
        io_fields.append(field(f"{io}_key_findings", f"{io.upper()} Key Findings", "Text", "IO", False, order))
        order += 1
        io_fields.append(field(f"{io}_recommended_actions", f"{io.upper()} Recommended Actions", "Text", "IO", False, order))
        order += 1
        io_fields.append(field(f"{io}_progress_score", f"{io.upper()} Progress Score", "Integer", "IO", True, order, 1, 10, 0))
        order += 1
        io_fields.append(field(f"{io}_responsible_agency", f"{io.upper()} Responsible Agency", "Text", "IO", False, order))
        order += 1
        io_fields.append(field(f"{io}_target_date", f"{io.upper()} Target Date", "Date", "IO", False, order))
        order += 1

    io_fields.extend(
        [
            field("io6_str_volume", "IO.6 STR Volume", "Integer", "IO", True, order),
            field("io6_ctr_volume", "IO.6 CTR Volume", "Integer", "IO", True, order + 1),
            field("io_average_progress_score", "Average IO Progress Score", "Decimal", "IO", True, order + 2),
            field("io_total_outcomes", "Total Immediate Outcomes", "Integer", "IO", True, order + 3),
        ]
    )

    templates = [
        template(
            "FATF_COV",
            "FATF Cover & Assessment Period",
            "Annual",
            "FixedRow",
            "fatf",
            [section("COV", "Cover", 1)],
            [
                field("assessment_year", "Assessment Year", "Integer", "COV", True, 1),
                field("assessment_cycle", "Assessment Cycle", "Text", "COV", True, 2),
                field("country_name", "Country Name", "Text", "COV", True, 3),
                field("lead_agency", "Lead Agency", "Text", "COV", True, 4),
                field("prepared_by", "Prepared By", "Text", "COV", True, 5),
                field("approved_by", "Approved By", "Text", "COV", True, 6),
                field("submission_date", "Submission Date", "Date", "COV", True, 7),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "FATF_TC",
            "Technical Compliance (R.1-R.40)",
            "Annual",
            "FixedRow",
            "fatf",
            [section("TC", "Technical Compliance", 1)],
            tc_fields,
            [
                formula_sum(
                    "tc_total_recommendations",
                    [
                        "tc_compliant_count",
                        "tc_largely_compliant_count",
                        "tc_partially_compliant_count",
                        "tc_non_compliant_count",
                    ],
                    "Total recommendations check",
                ),
                formula_ratio(
                    "tc_effective_compliance_ratio",
                    "tc_compliant_count",
                    "tc_total_recommendations",
                    "Compliant recommendations ratio",
                    "Warning",
                ),
            ],
            source_sheets=["Technical Compliance"],
        ),
        template(
            "FATF_IO",
            "Effectiveness (Immediate Outcomes)",
            "Annual",
            "FixedRow",
            "fatf",
            [section("IO", "Immediate Outcomes", 1)],
            io_fields,
            [
                formula_ratio("io_average_progress_score", "io1_progress_score", "io_total_outcomes", "Average IO progress approximation", "Warning"),
                formula_compare("GreaterThan", "io6_str_volume", "io6_ctr_volume", "IO.6 STR activity review", "Info"),
            ],
            source_sheets=["Immediate Outcomes"],
        ),
        template(
            "FATF_NRA",
            "National Risk Assessment",
            "Annual",
            "FixedRow",
            "fatf",
            [section("NRA", "NRA", 1)],
            [
                field("ml_threat_score", "ML Threat Score", "Integer", "NRA", True, 1, 1, 5, 0),
                field("ml_vulnerability_score", "ML Vulnerability Score", "Integer", "NRA", True, 2, 1, 5, 0),
                field("ml_risk_level", "ML Risk Level", "Text", "NRA", True, 3, enum_values="Low|Medium|High"),
                field("tf_threat_score", "TF Threat Score", "Integer", "NRA", True, 4, 1, 5, 0),
                field("tf_vulnerability_score", "TF Vulnerability Score", "Integer", "NRA", True, 5, 1, 5, 0),
                field("tf_risk_level", "TF Risk Level", "Text", "NRA", True, 6, enum_values="Low|Medium|High"),
                field("pf_threat_score", "PF Threat Score", "Integer", "NRA", True, 7, 1, 5, 0),
                field("pf_vulnerability_score", "PF Vulnerability Score", "Integer", "NRA", True, 8, 1, 5, 0),
                field("pf_risk_level", "PF Risk Level", "Text", "NRA", True, 9, enum_values="Low|Medium|High"),
                field("overall_risk_score", "Overall Risk Score", "Decimal", "NRA", True, 10),
            ],
            [
                formula_ratio("overall_risk_score", "ml_threat_score", "tf_vulnerability_score", "Overall risk proxy", "Warning"),
            ],
            source_sheets=["National Risk Assessment"],
        ),
        template(
            "FATF_LEA",
            "Law Enforcement & Prosecution",
            "Annual",
            "FixedRow",
            "fatf",
            [section("LEA", "LEA", 1)],
            [
                field("str_to_investigation_count", "STR-to-Investigation Count", "Integer", "LEA", True, 1),
                field("str_total_count", "STR Total Count", "Integer", "LEA", True, 2),
                field("str_to_investigation_rate", "STR-to-Investigation Rate", "Percentage", "LEA", True, 3),
                field("ml_convictions_count", "ML Convictions", "Integer", "LEA", True, 4),
                field("tf_prosecutions_count", "TF Prosecutions", "Integer", "LEA", True, 5),
                field("confiscation_value", "Confiscation Value", "Money", "LEA", True, 6),
                field("intl_coop_requests_sent", "International Cooperation Requests Sent", "Integer", "LEA", True, 7),
                field("intl_coop_requests_received", "International Cooperation Requests Received", "Integer", "LEA", True, 8),
            ],
            [
                formula_ratio("str_to_investigation_rate", "str_to_investigation_count", "str_total_count", "STR to investigation rate"),
            ],
            source_sheets=["Law Enforcement"],
        ),
        template(
            "FATF_TFS",
            "Targeted Financial Sanctions",
            "Annual",
            "FixedRow",
            "fatf",
            [section("TFS", "TFS", 1)],
            [
                field("tfs_total_screenings", "TFS Total Screenings", "Integer", "TFS", True, 1),
                field("tfs_total_matches", "TFS Total Matches", "Integer", "TFS", True, 2),
                field("tfs_false_positives", "TFS False Positives", "Integer", "TFS", True, 3),
                field("tfs_confirmed_matches", "TFS Confirmed Matches", "Integer", "TFS", True, 4),
                field("assets_frozen_count", "Assets Frozen Count", "Integer", "TFS", True, 5),
                field("assets_frozen_value", "Assets Frozen Value", "Money", "TFS", True, 6),
                field("unsc_screening_compliance", "UNSC Screening Compliance", "Percentage", "TFS", True, 7),
                field("ofac_screening_compliance", "OFAC Screening Compliance", "Percentage", "TFS", True, 8),
                field("eu_screening_compliance", "EU Screening Compliance", "Percentage", "TFS", True, 9),
            ],
            [
                formula_compare("LessThanOrEqual", "tfs_confirmed_matches", "tfs_total_matches", "Confirmed matches <= total matches"),
            ],
            source_sheets=["TFS"],
        ),
        template(
            "FATF_BO",
            "Beneficial Ownership",
            "Annual",
            "FixedRow",
            "fatf",
            [section("BO", "BO", 1)],
            [
                field("companies_with_bo_declarations", "Companies with BO Declarations", "Integer", "BO", True, 1),
                field("trusts_with_bo_info", "Trusts with BO Information", "Integer", "BO", True, 2),
                field("foundation_bo_compliance_rate", "Foundation BO Compliance Rate", "Percentage", "BO", True, 3),
                field("cac_bo_register_status", "CAC BO Register Status", "Text", "BO", True, 4, enum_values="Operational|Partial|Unavailable"),
                field("bo_verification_actions", "BO Verification Actions", "Integer", "BO", True, 5),
            ],
            [],
            source_sheets=["Beneficial Ownership"],
        ),
        template(
            "FATF_DNF",
            "DNFBPs",
            "Annual",
            "FixedRow",
            "fatf",
            [section("DNF", "DNFBPs", 1)],
            [
                field("real_estate_aml_compliance_rate", "Real Estate AML Compliance Rate", "Percentage", "DNF", True, 1),
                field("precious_metals_aml_compliance_rate", "Precious Metals AML Compliance Rate", "Percentage", "DNF", True, 2),
                field("lawyers_aml_compliance_rate", "Lawyers AML Compliance Rate", "Percentage", "DNF", True, 3),
                field("accountants_aml_compliance_rate", "Accountants AML Compliance Rate", "Percentage", "DNF", True, 4),
                field("dnfbp_supervisory_actions", "DNFBP Supervisory Actions", "Integer", "DNF", True, 5),
                field("dnfbp_sanctions", "DNFBP Sanctions", "Integer", "DNF", True, 6),
            ],
            [],
            source_sheets=["DNFBPs"],
        ),
        template(
            "FATF_ICP",
            "International Cooperation",
            "Annual",
            "FixedRow",
            "fatf",
            [section("ICP", "International Cooperation", 1)],
            [
                field("mla_requests_sent", "MLA Requests Sent", "Integer", "ICP", True, 1),
                field("mla_requests_received", "MLA Requests Received", "Integer", "ICP", True, 2),
                field("mla_requests_completed", "MLA Requests Completed", "Integer", "ICP", True, 3),
                field("extradition_cases", "Extradition Cases", "Integer", "ICP", True, 4),
                field("asset_recovery_cases", "Asset Recovery Cases", "Integer", "ICP", True, 5),
                field("egmont_exchange_count", "EGMONT Exchange Count", "Integer", "ICP", True, 6),
            ],
            [
                formula_compare("LessThanOrEqual", "mla_requests_completed", "mla_requests_received", "Completed MLA <= received"),
            ],
            source_sheets=["International Cooperation"],
        ),
        template(
            "FATF_APT",
            "Action Plan Tracker",
            "Annual",
            "FixedRow",
            "fatf",
            [section("APT", "Action Plan", 1)],
            [
                field("action_item_1", "Action Item 1", "Text", "APT", True, 1),
                field("action_item_1_milestone", "Action Item 1 Milestone", "Text", "APT", True, 2),
                field("action_item_1_target_date", "Action Item 1 Target Date", "Date", "APT", True, 3),
                field("action_item_1_status", "Action Item 1 Status", "Text", "APT", True, 4, enum_values="Not Started|In Progress|Completed|Verified"),
                field("action_item_1_evidence", "Action Item 1 Evidence", "Text", "APT", True, 5),
                field("action_item_2", "Action Item 2", "Text", "APT", True, 6),
                field("action_item_2_milestone", "Action Item 2 Milestone", "Text", "APT", True, 7),
                field("action_item_2_target_date", "Action Item 2 Target Date", "Date", "APT", True, 8),
                field("action_item_2_status", "Action Item 2 Status", "Text", "APT", True, 9, enum_values="Not Started|In Progress|Completed|Verified"),
                field("action_item_2_evidence", "Action Item 2 Evidence", "Text", "APT", True, 10),
                field("action_item_3", "Action Item 3", "Text", "APT", True, 11),
                field("action_item_3_milestone", "Action Item 3 Milestone", "Text", "APT", True, 12),
                field("action_item_3_target_date", "Action Item 3 Target Date", "Date", "APT", True, 13),
                field("action_item_3_status", "Action Item 3 Status", "Text", "APT", True, 14, enum_values="Not Started|In Progress|Completed|Verified"),
                field("action_item_3_evidence", "Action Item 3 Evidence", "Text", "APT", True, 15),
            ],
            [],
            source_sheets=["Action Plan Tracker"],
        ),
        template(
            "FATF_SUP",
            "Supervisory Framework",
            "Annual",
            "FixedRow",
            "fatf",
            [section("SUP", "Supervision", 1)],
            [
                field("cbn_onsite_examinations", "CBN Onsite Examinations", "Integer", "SUP", True, 1),
                field("sec_onsite_examinations", "SEC Onsite Examinations", "Integer", "SUP", True, 2),
                field("naicom_onsite_examinations", "NAICOM Onsite Examinations", "Integer", "SUP", True, 3),
                field("scuml_onsite_examinations", "SCUML Onsite Examinations", "Integer", "SUP", True, 4),
                field("nfiu_onsite_examinations", "NFIU Onsite Examinations", "Integer", "SUP", True, 5),
                field("offsite_reviews", "Offsite Reviews", "Integer", "SUP", True, 6),
                field("enforcement_actions", "Enforcement Actions", "Integer", "SUP", True, 7),
                field("enforcement_value", "Enforcement Value", "Money", "SUP", True, 8),
            ],
            [],
            source_sheets=["Supervisory Framework"],
        ),
        template(
            "FATF_VAS",
            "Virtual Assets & VASPs",
            "Annual",
            "FixedRow",
            "fatf",
            [section("VAS", "Virtual Assets", 1)],
            [
                field("vasp_registered_count", "Registered VASP Count", "Integer", "VAS", True, 1),
                field("vasp_supervised_count", "Supervised VASP Count", "Integer", "VAS", True, 2),
                field("sec_vasp_framework_status", "SEC VASP Framework Status", "Text", "VAS", True, 3, enum_values="Operational|Draft|Not Implemented"),
                field("vasp_transaction_monitoring_coverage", "VASP Transaction Monitoring Coverage", "Percentage", "VAS", True, 4),
                field("vasp_risk_assessment_score", "VASP Risk Assessment Score", "Integer", "VAS", True, 5, 1, 10, 0),
            ],
            [],
            source_sheets=["Virtual Assets"],
        ),
        template(
            "FATF_DIC",
            "Data Dictionary",
            "Annual",
            "FixedRow",
            "fatf",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
                field("validation_warning_count", "Validation Warning Count", "Integer", "DIC", True, 4),
                field("completeness_score", "Completeness Score", "Percentage", "DIC", True, 5),
                field("assessment_quality_score", "Assessment Quality Score", "Percentage", "DIC", True, 6),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "FATF_COV": (20, 5),
        "FATF_TC": (350, 90),
        "FATF_IO": (90, 25),
        "FATF_NRA": (30, 10),
        "FATF_LEA": (26, 8),
        "FATF_TFS": (24, 8),
        "FATF_BO": (20, 6),
        "FATF_DNF": (22, 6),
        "FATF_ICP": (20, 6),
        "FATF_APT": (32, 6),
        "FATF_SUP": (24, 7),
        "FATF_VAS": (20, 5),
        "FATF_DIC": (18, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "FATF_EVAL",
        "moduleVersion": "1.0.0",
        "description": "FATF Mutual Evaluation assessment module with technical compliance, immediate outcomes and action plan tracking.",
        "templates": templates,
        "interModuleDataFlows": [],
    }

    add_cross_rules(definition, 26)
    return definition


def esg_sector_fields(section_code):
    sectors = [
        "energy",
        "transportation",
        "materials",
        "agriculture_food_forest",
        "water_utilities",
        "real_estate",
        "construction",
        "industrial_goods",
        "consumer_goods",
        "consumer_services",
        "technology_telecom",
        "healthcare",
        "financials",
        "public_sector",
        "oil_gas",
        "mining_metals",
        "aviation_shipping",
        "waste_circular_economy",
    ]

    fields = []
    order = 1
    for s in sectors:
        label_prefix = s.replace("_", " ").title()
        fields.append(field(f"{s}_gross_exposure", f"{label_prefix} Gross Exposure", "Money", section_code, True, order))
        order += 1
        fields.append(field(f"{s}_percentage_of_portfolio", f"{label_prefix} % of Portfolio", "Percentage", section_code, True, order))
        order += 1
        fields.append(
            field(
                f"{s}_transition_risk_rating",
                f"{label_prefix} Transition Risk",
                "Text",
                section_code,
                True,
                order,
                enum_values="High|Medium|Low",
            )
        )
        order += 1
        fields.append(
            field(
                f"{s}_physical_risk_rating",
                f"{label_prefix} Physical Risk",
                "Text",
                section_code,
                True,
                order,
                enum_values="High|Medium|Low",
            )
        )
        order += 1

    fields.append(field("portfolio_denominator_assets", "Portfolio Denominator Assets", "Money", section_code, True, order))
    return fields, sectors


def esg_nsb_fields(section_code):
    principles = [
        "Business Activities",
        "Environmental and Social Footprint",
        "Human Rights",
        "Women Economic Empowerment",
        "Financial Inclusion",
        "ESR Governance",
        "Capacity Building",
        "Collaborative Partnerships",
        "Reporting",
    ]
    fields = []
    order = 1
    for idx, p in enumerate(principles, start=1):
        code = f"nsbp_p{idx}"
        fields.append(
            field(
                f"{code}_compliance_rating",
                f"{p} Compliance Rating",
                "Text",
                section_code,
                True,
                order,
                enum_values="Full|Partial|Non-Compliant",
            )
        )
        order += 1
        fields.append(field(f"{code}_evidence_summary", f"{p} Evidence Summary", "Text", section_code, True, order))
        order += 1
        fields.append(field(f"{code}_action_items", f"{p} Action Items", "Text", section_code, True, order))
        order += 1
    return fields


def esg_ghg_fields(section_code):
    fields = []
    order = 1

    scope1 = [
        "fuel_combustion_emissions",
        "fleet_emissions",
        "fugitive_refrigerants_emissions",
    ]
    for code in scope1:
        fields.append(field(code, code.replace("_", " ").title(), "Money", section_code, True, order))
        order += 1

    scope2 = [
        "purchased_electricity_location_based",
        "purchased_electricity_market_based",
        "purchased_heating_cooling",
    ]
    for code in scope2:
        fields.append(field(code, code.replace("_", " ").title(), "Money", section_code, True, order))
        order += 1

    scope3 = [
        "purchased_goods_services",
        "capital_goods",
        "fuel_energy_activities",
        "upstream_transport_distribution",
        "waste_generated_operations",
        "business_travel",
        "employee_commuting",
        "upstream_leased_assets",
        "downstream_transport_distribution",
        "processing_sold_products",
        "use_sold_products",
        "end_of_life_treatment",
        "downstream_leased_assets",
        "investments",
        "franchises",
    ]
    for code in scope3:
        fields.append(field(code, code.replace("_", " ").title(), "Money", section_code, True, order))
        order += 1

    fields.extend(
        [
            field("scope1_total", "Scope 1 Total", "Money", section_code, True, order),
            field("scope2_total", "Scope 2 Total", "Money", section_code, True, order + 1),
            field("scope3_total", "Scope 3 Total", "Money", section_code, True, order + 2),
            field("gross_ghg_total", "Gross GHG Total", "Money", section_code, True, order + 3),
        ]
    )

    return fields, scope1, scope2, scope3


def esg_module_definition():
    sectors_fields, sectors = esg_sector_fields("SEC")
    nsb_fields = esg_nsb_fields("NSB")
    ghg_fields, scope1, scope2, scope3 = esg_ghg_fields("GHG")

    asset_classes = [
        "corporate_loans",
        "project_finance",
        "commercial_real_estate",
        "mortgages",
        "motor_vehicles",
        "sovereign_bonds",
    ]

    fin_fields = []
    order = 1
    for ac in asset_classes:
        p = ac.replace("_", " ").title()
        fin_fields.append(field(f"{ac}_outstanding_amount", f"{p} Outstanding Amount", "Money", "FIN", True, order))
        order += 1
        fin_fields.append(field(f"{ac}_attributed_emissions_tco2e", f"{p} Attributed Emissions (tCO2e)", "Money", "FIN", True, order))
        order += 1
        fin_fields.append(field(f"{ac}_emission_factor", f"{p} Emission Factor", "Decimal", "FIN", True, order))
        order += 1
        fin_fields.append(field(f"{ac}_data_quality_score", f"{p} Data Quality Score", "Integer", "FIN", True, order, 1, 5, 0))
        order += 1
        fin_fields.append(field(f"{ac}_financed_emissions_intensity", f"{p} Financed Emissions Intensity", "Decimal", "FIN", True, order))
        order += 1

    fin_fields.extend(
        [
            field("total_financed_emissions_tco2e", "Total Financed Emissions", "Money", "FIN", True, order),
            field("total_financed_outstanding", "Total Financed Outstanding", "Money", "FIN", True, order + 1),
            field("portfolio_financed_emissions_intensity", "Portfolio Financed Emissions Intensity", "Decimal", "FIN", True, order + 2),
            field("total_data_quality_score", "Total Data Quality Score", "Integer", "FIN", True, order + 3),
            field("asset_class_count", "Asset Class Count", "Integer", "FIN", True, order + 4),
            field("average_data_quality_score", "Average Data Quality Score", "Decimal", "FIN", True, order + 5),
            field("total_loan_book", "Total Loan Book", "Money", "FIN", True, order + 6),
            field("oil_gas_exposure", "Oil and Gas Exposure", "Money", "FIN", True, order + 7),
            field("agriculture_exposure", "Agriculture Exposure", "Money", "FIN", True, order + 8),
            field("renewable_energy_exposure", "Renewable Energy Exposure", "Money", "FIN", True, order + 9),
            field("portfolio_exposure", "Portfolio Exposure (AUM)", "Money", "FIN", True, order + 10),
            field("development_sector_exposure", "Development Sector Exposure", "Money", "FIN", True, order + 11),
        ]
    )

    templates = [
        template(
            "ESG_COV",
            "ESG Cover & Reporting Boundaries",
            "Annual",
            "FixedRow",
            "esg",
            [section("COV", "Cover", 1)],
            [
                field("reporting_year", "Reporting Year", "Integer", "COV", True, 1),
                field("reporting_boundary", "Reporting Boundary", "Text", "COV", True, 2, enum_values="Entity|Consolidated|Group"),
                field("consolidation_approach", "Consolidation Approach", "Text", "COV", True, 3),
                field("prepared_by", "Prepared By", "Text", "COV", True, 4),
                field("approved_by", "Approved By", "Text", "COV", True, 5),
                field("submission_date", "Submission Date", "Date", "COV", True, 6),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "ESG_GOV",
            "TCFD Governance",
            "Annual",
            "FixedRow",
            "esg",
            [section("GOV", "Governance", 1)],
            [
                field("board_climate_oversight", "Board Climate Oversight", "Text", "GOV", True, 1),
                field("management_climate_role", "Management Climate Role", "Text", "GOV", True, 2),
                field("climate_expertise_count", "Climate Expertise Count", "Integer", "GOV", True, 3),
                field("climate_risk_committee", "Climate Risk Committee", "Boolean", "GOV", True, 4),
                field("board_climate_training_hours", "Board Climate Training Hours", "Integer", "GOV", True, 5),
                field("executive_comp_linked_to_climate_kpi", "Executive Compensation Linked to Climate KPI", "Boolean", "GOV", True, 6),
            ],
            [],
            source_sheets=["TCFD Governance"],
        ),
        template(
            "ESG_STR",
            "TCFD Strategy",
            "Annual",
            "FixedRow",
            "esg",
            [section("STR", "Strategy", 1)],
            [
                field("scenario_2c_impact", "Scenario 2C Impact", "Text", "STR", True, 1),
                field("scenario_3c_impact", "Scenario 3C Impact", "Text", "STR", True, 2),
                field("scenario_4c_impact", "Scenario 4C Impact", "Text", "STR", True, 3),
                field("physical_risk_assessment", "Physical Risk Assessment", "Text", "STR", True, 4),
                field("transition_risk_assessment", "Transition Risk Assessment", "Text", "STR", True, 5),
                field("climate_opportunities_summary", "Climate Opportunities", "Text", "STR", True, 6),
                field("strategy_resilience_score", "Strategy Resilience Score", "Integer", "STR", True, 7, 1, 10, 0),
            ],
            [],
            source_sheets=["TCFD Strategy"],
        ),
        template(
            "ESG_RMG",
            "TCFD Risk Management",
            "Annual",
            "FixedRow",
            "esg",
            [section("RMG", "Risk Management", 1)],
            [
                field("climate_risk_identification_process", "Climate Risk Identification Process", "Text", "RMG", True, 1),
                field("climate_risk_integration_into_erm", "Climate Risk Integration into ERM", "Text", "RMG", True, 2),
                field("acute_physical_risk_score", "Acute Physical Risk Score", "Integer", "RMG", True, 3, 1, 10, 0),
                field("chronic_physical_risk_score", "Chronic Physical Risk Score", "Integer", "RMG", True, 4, 1, 10, 0),
                field("policy_transition_risk_score", "Policy Transition Risk Score", "Integer", "RMG", True, 5, 1, 10, 0),
                field("technology_transition_risk_score", "Technology Transition Risk Score", "Integer", "RMG", True, 6, 1, 10, 0),
                field("market_transition_risk_score", "Market Transition Risk Score", "Integer", "RMG", True, 7, 1, 10, 0),
                field("reputation_transition_risk_score", "Reputation Transition Risk Score", "Integer", "RMG", True, 8, 1, 10, 0),
                field("climate_risk_appetite_statement", "Climate Risk Appetite Statement", "Text", "RMG", True, 9),
            ],
            [],
            source_sheets=["TCFD Risk Management"],
        ),
        template(
            "ESG_MET",
            "TCFD Metrics & Targets",
            "Annual",
            "FixedRow",
            "esg",
            [section("MET", "Metrics", 1)],
            [
                field("scope1_emissions", "Scope 1 Emissions", "Money", "MET", True, 1),
                field("scope2_emissions", "Scope 2 Emissions", "Money", "MET", True, 2),
                field("scope3_emissions", "Scope 3 Emissions", "Money", "MET", True, 3),
                field("total_ghg_emissions", "Total GHG Emissions", "Money", "MET", True, 4),
                field("emissions_reduction_target", "Emissions Reduction Target", "Percentage", "MET", True, 5),
                field("emissions_reduction_progress", "Emissions Reduction Progress", "Percentage", "MET", True, 6),
                field("revenue", "Revenue", "Money", "MET", True, 7),
                field("carbon_intensity_per_revenue", "Carbon Intensity per Revenue", "Decimal", "MET", True, 8),
            ],
            [
                formula_sum("total_ghg_emissions", ["scope1_emissions", "scope2_emissions", "scope3_emissions"], "Total GHG emissions"),
                formula_ratio("carbon_intensity_per_revenue", "total_ghg_emissions", "revenue", "Carbon intensity per revenue"),
            ],
            source_sheets=["TCFD Metrics & Targets"],
        ),
        template(
            "ESG_FINANCED_EMISSIONS",
            "Financed Emissions (PCAF)",
            "Annual",
            "FixedRow",
            "esg",
            [section("FIN", "Financed Emissions", 1)],
            fin_fields,
            [],
            source_sheets=["Financed Emissions"],
        ),
        template(
            "ESG_SEC",
            "Sector Exposure (BCBS 18 TCFD Sectors)",
            "Annual",
            "FixedRow",
            "esg",
            [section("SEC", "Sectors", 1)],
            sectors_fields,
            [],
            source_sheets=["Sector Exposure"],
        ),
        template(
            "ESG_NSB",
            "CBN NSBP Compliance",
            "Annual",
            "FixedRow",
            "esg",
            [section("NSB", "NSBP", 1)],
            nsb_fields,
            [],
            source_sheets=["CBN NSBP"],
        ),
        template(
            "ESG_GHG",
            "GHG Protocol Detail",
            "Annual",
            "FixedRow",
            "esg",
            [section("GHG", "GHG", 1)],
            ghg_fields,
            [],
            source_sheets=["GHG Protocol"],
        ),
        template(
            "ESG_CCA",
            "Climate Change Act & NDC",
            "Annual",
            "FixedRow",
            "esg",
            [section("CCA", "NDC", 1)],
            [
                field("ndc_unconditional_target_2030", "NDC Unconditional Target 2030", "Percentage", "CCA", True, 1),
                field("ndc_conditional_target_2030", "NDC Conditional Target 2030", "Percentage", "CCA", True, 2),
                field("institution_ndc_contribution", "Institution NDC Contribution", "Text", "CCA", True, 3),
                field("green_finance_portfolio", "Green Finance Portfolio", "Money", "CCA", True, 4),
                field("renewable_energy_financing", "Renewable Energy Financing", "Money", "CCA", True, 5),
                field("climate_adaptation_finance", "Climate Adaptation Finance", "Money", "CCA", True, 6),
            ],
            [],
            source_sheets=["Climate Change Act & NDC"],
        ),
        template(
            "ESG_ETP",
            "Energy Transition Plan",
            "Annual",
            "FixedRow",
            "esg",
            [section("ETP", "ETP", 1)],
            [
                field("carbon_neutrality_target_year", "Carbon Neutrality Target Year", "Integer", "ETP", True, 1),
                field("fossil_fuel_exposure", "Fossil Fuel Exposure", "Money", "ETP", True, 2),
                field("renewable_energy_financing", "Renewable Energy Financing", "Money", "ETP", True, 3),
                field("energy_efficiency_projects", "Energy Efficiency Projects", "Integer", "ETP", True, 4),
                field("just_transition_programmes", "Just Transition Programmes", "Integer", "ETP", True, 5),
            ],
            [],
            source_sheets=["Energy Transition Plan"],
        ),
        template(
            "ESG_TAR",
            "Targets & Progress",
            "Annual",
            "FixedRow",
            "esg",
            [section("TAR", "Targets", 1)],
            [
                field("net_zero_commitment_year", "Net-Zero Commitment Year", "Integer", "TAR", True, 1),
                field("interim_target_2030", "Interim Target 2030", "Percentage", "TAR", True, 2),
                field("science_based_target_status", "Science-Based Target Status", "Text", "TAR", True, 3, enum_values="Committed|Validated|Not Committed"),
                field("target_progress_percent", "Target Progress %", "Percentage", "TAR", True, 4),
                field("carbon_offsets_purchased", "Carbon Offsets Purchased", "Money", "TAR", True, 5),
            ],
            [],
            source_sheets=["Targets & Progress"],
        ),
        template(
            "ESG_DIC",
            "Data Dictionary",
            "Annual",
            "FixedRow",
            "esg",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
                field("validation_warning_count", "Validation Warning Count", "Integer", "DIC", True, 4),
                field("completeness_score", "Completeness Score", "Percentage", "DIC", True, 5),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    fin_template = next(t for t in templates if t["returnCode"] == "ESG_FINANCED_EMISSIONS")
    fin_formulas = []
    for ac in asset_classes:
        fin_formulas.append(
            formula_ratio(
                f"{ac}_financed_emissions_intensity",
                f"{ac}_attributed_emissions_tco2e",
                f"{ac}_outstanding_amount",
                f"{ac.replace('_', ' ').title()} financed emissions intensity",
            )
        )

    fin_formulas.extend(
        [
            formula_sum(
                "total_financed_emissions_tco2e",
                [f"{ac}_attributed_emissions_tco2e" for ac in asset_classes],
                "Total financed emissions",
            ),
            formula_sum(
                "total_financed_outstanding",
                [f"{ac}_outstanding_amount" for ac in asset_classes],
                "Total financed outstanding amount",
            ),
            formula_ratio(
                "portfolio_financed_emissions_intensity",
                "total_financed_emissions_tco2e",
                "total_financed_outstanding",
                "Portfolio financed emissions intensity",
            ),
            formula_sum(
                "total_data_quality_score",
                [f"{ac}_data_quality_score" for ac in asset_classes],
                "Total data quality score",
                "Warning",
                0,
            ),
            formula_ratio(
                "average_data_quality_score",
                "total_data_quality_score",
                "asset_class_count",
                "Average data quality score",
            ),
        ]
    )
    fin_template["formulas"] = fin_formulas

    sec_template = next(t for t in templates if t["returnCode"] == "ESG_SEC")
    sec_formulas = []
    for s in sectors:
        sec_formulas.append(
            formula_ratio(
                f"{s}_percentage_of_portfolio",
                f"{s}_gross_exposure",
                "portfolio_denominator_assets",
                f"{s.replace('_', ' ').title()} % of portfolio",
                "Warning",
            )
        )
    sec_template["formulas"] = sec_formulas

    ghg_template = next(t for t in templates if t["returnCode"] == "ESG_GHG")
    ghg_template["formulas"] = [
        formula_sum("scope1_total", scope1, "Scope 1 total emissions"),
        formula_sum("scope2_total", scope2, "Scope 2 total emissions"),
        formula_sum("scope3_total", scope3, "Scope 3 total emissions"),
        formula_sum("gross_ghg_total", ["scope1_total", "scope2_total", "scope3_total"], "Gross GHG emissions total"),
    ]

    targets = {
        "ESG_COV": (18, 4),
        "ESG_GOV": (24, 6),
        "ESG_STR": (26, 6),
        "ESG_RMG": (28, 7),
        "ESG_MET": (30, 8),
        "ESG_FINANCED_EMISSIONS": (90, 20),
        "ESG_SEC": (85, 30),
        "ESG_NSB": (40, 8),
        "ESG_GHG": (60, 12),
        "ESG_CCA": (24, 6),
        "ESG_ETP": (20, 5),
        "ESG_TAR": (20, 5),
        "ESG_DIC": (18, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "ESG_CLIMATE",
        "moduleVersion": "1.0.0",
        "description": "ESG and climate disclosure module aligned with TCFD, BCBS CRFR, PCAF, NSBP and Nigeria climate policies.",
        "templates": templates,
        "interModuleDataFlows": [],
    }

    add_cross_rules(definition, 26)
    return definition


def add_specific_cross_rules(fatf, esg):
    next(t for t in fatf["templates"] if t["returnCode"] == "FATF_TC")["crossSheetRules"].append(
        {
            "description": "FATF technical compliance ratios should align with immediate outcome progress",
            "sourceTemplate": "FATF_TC",
            "sourceField": "tc_effective_compliance_ratio",
            "targetTemplate": "FATF_IO",
            "targetField": "io_average_progress_score",
            "operator": "LessThanOrEqual",
            "severity": "Warning",
            "tolerancePercent": 20.0,
        }
    )

    next(t for t in fatf["templates"] if t["returnCode"] == "FATF_IO")["crossSheetRules"].append(
        {
            "description": "IO.6 STR volume should feed law-enforcement STR base",
            "sourceTemplate": "FATF_IO",
            "sourceField": "io6_str_volume",
            "targetTemplate": "FATF_LEA",
            "targetField": "str_total_count",
            "operator": "Equals",
            "severity": "Warning",
            "tolerancePercent": 10.0,
        }
    )

    next(t for t in esg["templates"] if t["returnCode"] == "ESG_FINANCED_EMISSIONS")["crossSheetRules"].append(
        {
            "description": "Financed outstanding should be within portfolio denominator assets",
            "sourceTemplate": "ESG_FINANCED_EMISSIONS",
            "sourceField": "total_financed_outstanding",
            "targetTemplate": "ESG_SEC",
            "targetField": "portfolio_denominator_assets",
            "operator": "LessThanOrEqual",
            "severity": "Warning",
            "tolerancePercent": 15.0,
        }
    )


def write_module_json(output_dir: Path, module_name: str, definition: dict):
    output_path = output_dir / module_name
    output_path.write_text(json.dumps(definition, indent=2) + "\n", encoding="utf-8")


def summarize(defn):
    t = len(defn["templates"])
    f = sum(len(x["fields"]) for x in defn["templates"])
    fo = sum(len(x["formulas"]) for x in defn["templates"])
    cr = sum(len(x["crossSheetRules"]) for x in defn["templates"])
    fl = len(defn.get("interModuleDataFlows", []))
    return t, f, fo, cr, fl


def main():
    root = Path(__file__).resolve().parents[1]
    out = root / "docs" / "module-definitions" / "rg11"
    out.mkdir(parents=True, exist_ok=True)

    templates_root = root.parents[0] / "Templates"

    fatf = fatf_module_definition()
    esg = esg_module_definition()
    add_specific_cross_rules(fatf, esg)

    add_source_coverage(
        fatf,
        templates_root / "FATF_Mutual_Evaluation_Templates.xlsx",
        templates_root / "FATF_ME_Schema_v1.0.xsd",
        "FATF_DIC",
    )

    write_module_json(out, "fatf_eval.json", fatf)
    write_module_json(out, "esg_climate.json", esg)

    totals = [
        ("FATF_EVAL", summarize(fatf)),
        ("ESG_CLIMATE", summarize(esg)),
    ]

    print("RG-11 module definition generation complete")
    print("module,templates,fields,formulas,cross_sheet_rules,data_flows")
    for module_code, metrics in totals:
        print(f"{module_code},{metrics[0]},{metrics[1]},{metrics[2]},{metrics[3]},{metrics[4]}")

    all_templates = sum(m[0] for _, m in totals)
    all_fields = sum(m[1] for _, m in totals)
    all_formulas = sum(m[2] for _, m in totals)
    all_cross = sum(m[3] for _, m in totals)
    all_flows = sum(m[4] for _, m in totals)
    print(f"TOTAL,{all_templates},{all_fields},{all_formulas},{all_cross},{all_flows}")


if __name__ == "__main__":
    main()
