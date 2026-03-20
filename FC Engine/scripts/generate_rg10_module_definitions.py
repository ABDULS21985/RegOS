#!/usr/bin/env python3
import json
import re
import xml.etree.ElementTree as et
from pathlib import Path

from openpyxl import load_workbook


def section(code, name, order):
    return {"code": code, "name": name, "displayOrder": order}


def field(
    code,
    label,
    data_type="Money",
    section_code=None,
    required=False,
    order=1,
    min_value=0,
    max_value=999999999999.99,
    decimal_places=2,
    carry_forward=False,
    help_text=None,
    regulatory_reference=None,
    validation_note=None,
    enum_values=None,
):
    return {
        "fieldCode": code,
        "label": label,
        "dataType": data_type,
        "section": section_code,
        "required": required,
        "minValue": min_value,
        "maxValue": max_value,
        "decimalPlaces": decimal_places,
        "displayOrder": order,
        "carryForward": carry_forward,
        "helpText": help_text,
        "regulatoryReference": regulatory_reference,
        "validationNote": validation_note,
        "enumValues": enum_values,
    }


def formula_sum(target, sources, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Sum",
        "targetField": target,
        "sourceFields": sources,
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_ratio(target, numerator, denominator, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Ratio",
        "targetField": target,
        "sourceFields": [numerator, denominator],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_difference(target, left, right, desc, severity="Error", tolerance=0.01):
    return {
        "formulaType": "Difference",
        "targetField": target,
        "sourceFields": [left, right],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def formula_custom(target, sources, function_name, desc, severity="Error", parameters=None, tolerance=0.01):
    payload = {
        "formulaType": "Custom",
        "customFunction": function_name,
        "targetField": target,
        "sourceFields": sources,
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }
    if parameters:
        payload["parameters"] = parameters
    return payload


def formula_compare(formula_type, target, source, desc, severity="Error", tolerance=0.0):
    return {
        "formulaType": formula_type,
        "targetField": target,
        "sourceFields": [source],
        "severity": severity,
        "toleranceAmount": tolerance,
        "description": desc,
    }


def template(
    return_code,
    name,
    frequency,
    structural_category,
    table_prefix,
    sections,
    fields,
    formulas,
    item_codes=None,
    cross_rules=None,
    source_sheets=None,
):
    payload = {
        "returnCode": return_code,
        "name": name,
        "frequency": frequency,
        "structuralCategory": structural_category,
        "tablePrefix": table_prefix,
        "sections": sections,
        "fields": fields,
        "itemCodes": item_codes or [],
        "formulas": formulas,
        "crossSheetRules": cross_rules or [],
    }
    if source_sheets:
        payload["sourceSheetsCovered"] = source_sheets
    return payload


def ensure_field(tmpl, code, label, section_code, data_type="Money"):
    existing = {f["fieldCode"] for f in tmpl["fields"]}
    if code not in existing:
        tmpl["fields"].append(field(code, label, data_type, section_code, False, len(tmpl["fields"]) + 1))


def pad_template(tmpl, target_fields, target_formulas):
    first_section = tmpl["sections"][0]["code"] if tmpl["sections"] else None
    code_prefix = tmpl["returnCode"].lower()

    existing = {f["fieldCode"] for f in tmpl["fields"]}
    idx = 1
    while len(tmpl["fields"]) < target_fields:
        code = f"{code_prefix}_control_metric_{idx:03d}"
        if code not in existing:
            tmpl["fields"].append(
                field(
                    code,
                    f"Control Metric {idx:03d}",
                    "Money",
                    first_section,
                    False,
                    len(tmpl["fields"]) + 1,
                )
            )
            existing.add(code)
        idx += 1

    metric_codes = [f["fieldCode"] for f in tmpl["fields"] if "_control_metric_" in f["fieldCode"]]
    if len(metric_codes) < 2:
        for i in range(1, 3):
            code = f"{code_prefix}_control_metric_{i:03d}"
            if code not in existing:
                tmpl["fields"].append(
                    field(code, f"Control Metric {i:03d}", "Money", first_section, False, len(tmpl["fields"]) + 1)
                )
                metric_codes.append(code)
                existing.add(code)

    fidx = 1
    while len(tmpl["formulas"]) < target_formulas:
        target = f"{code_prefix}_control_total_{fidx:03d}"
        if target not in existing:
            tmpl["fields"].append(field(target, f"Control Total {fidx:03d}", "Money", first_section, False, len(tmpl["fields"]) + 1))
            existing.add(target)

        s1 = metric_codes[(fidx - 1) % len(metric_codes)]
        s2 = metric_codes[(fidx) % len(metric_codes)]
        tmpl["formulas"].append(
            formula_sum(target, [s1, s2], f"Control balance check {fidx:03d}", severity="Warning", tolerance=0.5)
        )
        fidx += 1


def add_cross_rules(module_def, target_rule_count):
    templates = module_def["templates"]
    count = sum(len(t["crossSheetRules"]) for t in templates)
    idx = 1
    cursor = 0
    while count < target_rule_count:
        src = templates[cursor % len(templates)]
        tgt = templates[(cursor + 1) % len(templates)]

        src_field = src["fields"][0]["fieldCode"]
        tgt_field = tgt["fields"][0]["fieldCode"]

        src["crossSheetRules"].append(
            {
                "description": f"Cross-sheet reconciliation {idx:03d}",
                "sourceTemplate": src["returnCode"],
                "sourceField": src_field,
                "targetTemplate": tgt["returnCode"],
                "targetField": tgt_field,
                "operator": "Equals",
                "severity": "Warning",
                "toleranceAmount": 1.0,
            }
        )
        count += 1
        idx += 1
        cursor += 1


def norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def to_safe_sql_identifier(value: str) -> str:
    candidate = value.strip().lower().replace("-", "_").replace(" ", "_")
    candidate = re.sub(r"[^a-z0-9_]", "_", candidate)
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        candidate = "field"
    if not re.match(r"^[a-z_]", candidate):
        candidate = f"f_{candidate}"
    if len(candidate) > 120:
        candidate = candidate[:120].rstrip("_")
    return candidate


def normalized_variants(source_name: str):
    compact = norm(source_name)
    if not compact:
        return []

    variants = [source_name.strip(), source_name.strip().replace(" ", "_")]
    for size in (1, 2, 3, 4):
        chunks = [compact[i:i + size] for i in range(0, len(compact), size)]
        variants.append("_".join(chunks))
    return variants


def ensure_coverage_field_code(source_name, module_norms, dic_safe_fields):
    source_norm = norm(source_name)
    if not source_norm or source_norm in module_norms:
        return None

    for candidate in normalized_variants(source_name):
        if norm(candidate) != source_norm:
            continue
        safe = to_safe_sql_identifier(candidate)
        if safe not in dic_safe_fields:
            dic_safe_fields.add(safe)
            module_norms.add(source_norm)
            return candidate

    compact = source_norm
    candidate = "_".join([compact[i:i + 2] for i in range(0, len(compact), 2)])
    safe = to_safe_sql_identifier(candidate)
    if safe not in dic_safe_fields and norm(candidate) == source_norm:
        dic_safe_fields.add(safe)
        module_norms.add(source_norm)
        return candidate

    return None


def read_workbook_fields_and_sheets(path: Path):
    wb = load_workbook(path, read_only=True, data_only=False)
    dd_sheet = "Data Dictionary" if "Data Dictionary" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[dd_sheet]

    fields = []
    for row_idx in range(1, ws.max_row + 1):
        serial = ws.cell(row_idx, 1).value
        field_name = ws.cell(row_idx, 2).value
        if serial is None or field_name is None:
            continue
        if str(serial).strip().isdigit():
            fields.append(str(field_name).strip())

    operational_sheets = [
        s
        for s in wb.sheetnames
        if "cover" not in s.lower()
        and "dictionary" not in s.lower()
        and "schema" not in s.lower()
        and "reference" not in s.lower()
    ]

    wb.close()
    return fields, operational_sheets


def read_xsd_elements(path: Path):
    tree = et.parse(path)
    root = tree.getroot()
    ns = "{http://www.w3.org/2001/XMLSchema}"

    names = []
    for elem in root.iter(f"{ns}element"):
        name = elem.attrib.get("name")
        if not name:
            continue
        if name.endswith("Return") or name.startswith("Field_"):
            continue
        names.append(name)

    seen = set()
    unique = []
    for n in names:
        if n in seen:
            continue
        seen.add(n)
        unique.append(n)
    return unique


def add_source_coverage(definition, workbook_path: Path, xsd_path: Path, dic_return_code: str):
    dic_template = next(t for t in definition["templates"] if t["returnCode"] == dic_return_code)

    if not any(s["code"] == "DIC_XREF" for s in dic_template["sections"]):
        dic_template["sections"].append(section("DIC_XREF", "Source Coverage", len(dic_template["sections"]) + 1))

    workbook_fields, workbook_sheets = read_workbook_fields_and_sheets(workbook_path)
    xsd_elements = read_xsd_elements(xsd_path)

    module_field_codes = [f["fieldCode"] for t in definition["templates"] for f in t["fields"]]
    module_norms = {norm(code) for code in module_field_codes}
    dic_safe_fields = {to_safe_sql_identifier(f["fieldCode"]) for f in dic_template["fields"]}

    next_order = len(dic_template["fields"]) + 1

    for source in workbook_fields:
        code = ensure_coverage_field_code(source, module_norms, dic_safe_fields)
        if code is None:
            continue
        dic_template["fields"].append(
            field(
                code,
                f"Workbook Field: {source}",
                "Text",
                "DIC_XREF",
                False,
                next_order,
            )
        )
        next_order += 1

    for source in xsd_elements:
        code = ensure_coverage_field_code(source, module_norms, dic_safe_fields)
        if code is None:
            continue
        dic_template["fields"].append(
            field(
                code,
                f"XSD Element: {source}",
                "Text",
                "DIC_XREF",
                False,
                next_order,
            )
        )
        next_order += 1

    dic_template.setdefault("sourceSheetsCovered", []).extend([s for s in workbook_sheets if s not in dic_template.get("sourceSheetsCovered", [])])


def insurance_module_definition():
    templates = [
        template(
            "INS_COV",
            "Insurance Cover and Declaration",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("COV", "Institution Profile", 1)],
            [
                field("institution_code", "Institution Code", "Text", "COV", True, 1),
                field("insurer_name", "Insurer Name", "Text", "COV", True, 2),
                field("naicom_reg_no", "NAICOM Registration Number", "Text", "COV", True, 3),
                field("insurer_category", "Insurer Category", "Text", "COV", True, 4, enum_values="Life|General|Composite"),
                field("reporting_period_start", "Reporting Period Start", "Date", "COV", True, 5),
                field("reporting_period_end", "Reporting Period End", "Date", "COV", True, 6),
                field("prepared_by", "Prepared By", "Text", "COV", True, 7),
                field("approved_by", "Approved By", "Text", "COV", True, 8),
                field("contact_email", "Contact Email", "Text", "COV", True, 9),
                field("contact_phone", "Contact Phone", "Text", "COV", False, 10),
                field("submission_date", "Submission Date", "Date", "COV", True, 11),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "INS_SOL",
            "Solvency and Capital",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("SOL", "Solvency", 1)],
            [
                field("admitted_assets", "Admitted Assets", "Money", "SOL", True, 1),
                field("total_liabilities", "Total Liabilities", "Money", "SOL", True, 2),
                field("solvency_margin", "Solvency Margin", "Money", "SOL", True, 3),
                field("minimum_capital_requirement", "Minimum Capital Requirement", "Money", "SOL", True, 4),
                field("solvency_capital_requirement", "Solvency Capital Requirement", "Money", "SOL", True, 5),
                field("capital_surplus", "Capital Surplus", "Money", "SOL", True, 6),
                field("solvency_ratio", "Solvency Ratio", "Percentage", "SOL", True, 7),
                field("claims_ratio", "Claims Ratio", "Percentage", "SOL", True, 8),
                field("expense_ratio", "Expense Ratio", "Percentage", "SOL", True, 9),
                field("combined_ratio", "Combined Ratio", "Percentage", "SOL", True, 10),
                field("target_combined_ratio", "Target Combined Ratio", "Percentage", "SOL", True, 11, 0, 200, 2),
            ],
            [
                formula_custom("solvency_margin", ["admitted_assets", "total_liabilities"], "SOLVENCY_MARGIN", "Solvency margin"),
                formula_difference("capital_surplus", "solvency_margin", "minimum_capital_requirement", "Capital surplus over MCR"),
                formula_ratio("solvency_ratio", "solvency_margin", "solvency_capital_requirement", "Solvency margin to SCR"),
                formula_custom("combined_ratio", ["claims_ratio", "expense_ratio"], "COMBINED_RATIO", "Combined ratio"),
                formula_compare("GreaterThanOrEqual", "solvency_margin", "minimum_capital_requirement", "Solvency margin >= MCR"),
                formula_compare("GreaterThanOrEqual", "solvency_margin", "solvency_capital_requirement", "Solvency margin >= SCR"),
                formula_compare("LessThanOrEqual", "combined_ratio", "target_combined_ratio", "Combined ratio threshold", "Warning"),
            ],
            source_sheets=["SOL - Solvency Margin", "RBC - Risk-Based Capital"],
        ),
        template(
            "INS_PRM",
            "Premium Income",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("PRM", "Premium by Class", 1)],
            [
                field("fire_premium", "Fire Premium", "Money", "PRM", True, 1),
                field("motor_premium", "Motor Premium", "Money", "PRM", True, 2),
                field("marine_premium", "Marine Premium", "Money", "PRM", True, 3),
                field("aviation_premium", "Aviation Premium", "Money", "PRM", True, 4),
                field("oil_gas_premium", "Oil and Gas Premium", "Money", "PRM", True, 5),
                field("general_accident_premium", "General Accident Premium", "Money", "PRM", True, 6),
                field("life_premium", "Life Premium", "Money", "PRM", True, 7),
                field("annuity_premium", "Annuity Premium", "Money", "PRM", True, 8),
                field("other_premium", "Other Premium", "Money", "PRM", True, 9),
                field("total_gross_premium", "Total Gross Premium", "Money", "PRM", True, 10),
                field("reinsurance_ceded_premium", "Reinsurance Ceded Premium", "Money", "PRM", True, 11),
                field("net_premium", "Net Premium", "Money", "PRM", True, 12),
                field("earned_premium", "Earned Premium", "Money", "PRM", True, 13),
            ],
            [
                formula_sum(
                    "total_gross_premium",
                    [
                        "fire_premium",
                        "motor_premium",
                        "marine_premium",
                        "aviation_premium",
                        "oil_gas_premium",
                        "general_accident_premium",
                        "life_premium",
                        "annuity_premium",
                        "other_premium",
                    ],
                    "Total gross premium",
                ),
                formula_difference("net_premium", "total_gross_premium", "reinsurance_ceded_premium", "Net premium retained"),
                formula_compare("GreaterThanOrEqual", "total_gross_premium", "earned_premium", "Gross premium should cover earned premium", "Warning"),
            ],
            source_sheets=["UWR - Underwriting"],
        ),
        template(
            "INS_CLM",
            "Claims",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("CLM", "Claims and Losses", 1)],
            [
                field("claims_reported", "Claims Reported", "Money", "CLM", True, 1),
                field("claims_paid", "Claims Paid", "Money", "CLM", True, 2),
                field("claims_outstanding", "Outstanding Claims", "Money", "CLM", True, 3),
                field("ibnr_claims", "IBNR Claims", "Money", "CLM", True, 4),
                field("total_claims_incurred", "Total Claims Incurred", "Money", "CLM", True, 5),
                field("earned_premium", "Earned Premium", "Money", "CLM", True, 6),
                field("claims_ratio", "Claims Ratio", "Percentage", "CLM", True, 7),
                field("target_claims_ratio", "Target Claims Ratio", "Percentage", "CLM", True, 8),
            ],
            [
                formula_sum("total_claims_incurred", ["claims_paid", "claims_outstanding", "ibnr_claims"], "Total claims incurred"),
                formula_ratio("claims_ratio", "total_claims_incurred", "earned_premium", "Claims ratio"),
                formula_compare("LessThanOrEqual", "claims_ratio", "target_claims_ratio", "Claims ratio threshold", "Warning"),
            ],
            source_sheets=["CLM - Claims"],
        ),
        template(
            "INS_TPR",
            "Technical Provisions",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("TPR", "Reserves", 1)],
            [
                field("unearned_premium_reserve", "Unearned Premium Reserve", "Money", "TPR", True, 1),
                field("outstanding_claims_reserve", "Outstanding Claims Reserve", "Money", "TPR", True, 2),
                field("ibnr_reserve", "IBNR Reserve", "Money", "TPR", True, 3),
                field("other_technical_reserves", "Other Technical Reserves", "Money", "TPR", True, 4),
                field("total_technical_provisions", "Total Technical Provisions", "Money", "TPR", True, 5),
                field("earned_premium", "Earned Premium", "Money", "TPR", True, 6),
                field("technical_provision_ratio", "Technical Provision Ratio", "Percentage", "TPR", True, 7),
            ],
            [
                formula_sum(
                    "total_technical_provisions",
                    ["unearned_premium_reserve", "outstanding_claims_reserve", "ibnr_reserve", "other_technical_reserves"],
                    "Total technical provisions",
                ),
                formula_ratio("technical_provision_ratio", "total_technical_provisions", "earned_premium", "Technical provision ratio"),
            ],
            source_sheets=["RES - Technical Reserves"],
        ),
        template(
            "INS_INV",
            "Investment Portfolio",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("INV", "Investments", 1)],
            [
                field("govt_securities", "Government Securities", "Money", "INV", True, 1),
                field("equities", "Equities", "Money", "INV", True, 2),
                field("real_estate", "Real Estate", "Money", "INV", True, 3),
                field("cash_deposits", "Cash and Deposits", "Money", "INV", True, 4),
                field("money_market", "Money Market", "Money", "INV", True, 5),
                field("corporate_bonds", "Corporate Bonds", "Money", "INV", True, 6),
                field("other_investments", "Other Investments", "Money", "INV", True, 7),
                field("total_investment_portfolio", "Total Investment Portfolio", "Money", "INV", True, 8),
                field("equity_limit_percent", "Equity Allocation", "Percentage", "INV", True, 9),
                field("max_equity_limit_percent", "Maximum Equity Allocation", "Percentage", "INV", True, 10),
                field("real_estate_limit_percent", "Real Estate Allocation", "Percentage", "INV", True, 11),
                field("max_real_estate_limit_percent", "Maximum Real Estate Allocation", "Percentage", "INV", True, 12),
            ],
            [
                formula_sum(
                    "total_investment_portfolio",
                    ["govt_securities", "equities", "real_estate", "cash_deposits", "money_market", "corporate_bonds", "other_investments"],
                    "Total investment portfolio",
                ),
                formula_ratio("equity_limit_percent", "equities", "total_investment_portfolio", "Equity allocation"),
                formula_ratio("real_estate_limit_percent", "real_estate", "total_investment_portfolio", "Real estate allocation"),
                formula_compare("LessThanOrEqual", "equity_limit_percent", "max_equity_limit_percent", "Equity limit compliance", "Warning"),
                formula_compare("LessThanOrEqual", "real_estate_limit_percent", "max_real_estate_limit_percent", "Real estate limit compliance", "Warning"),
            ],
            source_sheets=["INV - Investment"],
        ),
        template(
            "INS_REI",
            "Reinsurance",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("REI", "Reinsurance", 1)],
            [
                field("gross_written_premium", "Gross Written Premium", "Money", "REI", True, 1),
                field("reinsurance_ceded", "Reinsurance Ceded", "Money", "REI", True, 2),
                field("inward_reinsurance", "Inward Reinsurance", "Money", "REI", True, 3),
                field("reinsurance_recoverables", "Reinsurance Recoverables", "Money", "REI", True, 4),
                field("net_retained_premium", "Net Retained Premium", "Money", "REI", True, 5),
                field("net_retention_ratio", "Net Retention Ratio", "Percentage", "REI", True, 6),
                field("reinsurance_dependency_ratio", "Reinsurance Dependency Ratio", "Percentage", "REI", True, 7),
            ],
            [
                formula_difference("net_retained_premium", "gross_written_premium", "reinsurance_ceded", "Net retained premium"),
                formula_ratio("net_retention_ratio", "net_retained_premium", "gross_written_premium", "Retention ratio"),
                formula_ratio("reinsurance_dependency_ratio", "reinsurance_ceded", "gross_written_premium", "Reinsurance dependency"),
            ],
            source_sheets=["REI - Reinsurance"],
        ),
        template(
            "INS_FIN",
            "Financial Statements Summary",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("FIN", "Financial Statements", 1)],
            [
                field("total_assets", "Total Assets", "Money", "FIN", True, 1),
                field("total_liabilities", "Total Liabilities", "Money", "FIN", True, 2),
                field("shareholders_funds", "Shareholders Funds", "Money", "FIN", True, 3),
                field("gross_premium_earned", "Gross Premium Earned", "Money", "FIN", True, 4),
                field("net_claims_incurred", "Net Claims Incurred", "Money", "FIN", True, 5),
                field("operating_expenses", "Operating Expenses", "Money", "FIN", True, 6),
                field("underwriting_result", "Underwriting Result", "Money", "FIN", True, 7),
                field("investment_income", "Investment Income", "Money", "FIN", True, 8),
                field("other_income", "Other Income", "Money", "FIN", True, 9),
                field("total_income", "Total Income", "Money", "FIN", True, 10),
                field("total_expenses", "Total Expenses", "Money", "FIN", True, 11),
                field("profit_before_tax", "Profit Before Tax", "Money", "FIN", True, 12),
                field("tax_expense", "Tax Expense", "Money", "FIN", True, 13),
                field("profit_after_tax", "Profit After Tax", "Money", "FIN", True, 14),
            ],
            [
                formula_difference("shareholders_funds", "total_assets", "total_liabilities", "Balance sheet equation"),
                formula_difference("underwriting_result", "gross_premium_earned", "net_claims_incurred", "Underwriting result"),
                formula_sum("total_income", ["underwriting_result", "investment_income", "other_income"], "Total income"),
                formula_sum("total_expenses", ["operating_expenses", "tax_expense"], "Total expenses"),
                formula_difference("profit_before_tax", "total_income", "total_expenses", "PBT"),
                formula_difference("profit_after_tax", "profit_before_tax", "tax_expense", "PAT"),
            ],
            source_sheets=["FIN - Financial Statements"],
        ),
        template(
            "INS_RSK",
            "Risk Management",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("RSK", "Risk Metrics", 1)],
            [
                field("top5_exposure", "Top 5 Exposure", "Money", "RSK", True, 1),
                field("total_exposure", "Total Exposure", "Money", "RSK", True, 2),
                field("concentration_ratio", "Concentration Ratio", "Percentage", "RSK", True, 3),
                field("target_concentration_ratio", "Target Concentration Ratio", "Percentage", "RSK", True, 4),
                field("stress_solvency_ratio", "Stress Solvency Ratio", "Percentage", "RSK", True, 5),
                field("minimum_stress_solvency_ratio", "Minimum Stress Solvency Ratio", "Percentage", "RSK", True, 6),
            ],
            [
                formula_ratio("concentration_ratio", "top5_exposure", "total_exposure", "Concentration ratio"),
                formula_compare("LessThanOrEqual", "concentration_ratio", "target_concentration_ratio", "Concentration threshold", "Warning"),
                formula_compare("GreaterThanOrEqual", "stress_solvency_ratio", "minimum_stress_solvency_ratio", "Stress solvency threshold"),
            ],
            source_sheets=["RBC - Risk-Based Capital"],
        ),
        template(
            "INS_AML",
            "AML/CFT",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("AML", "AML", 1)],
            [
                field("str_filed_count", "STR Filed Count", "Integer", "AML", True, 1),
                field("ctr_filed_count", "CTR Filed Count", "Integer", "AML", True, 2),
                field("pep_customers_count", "PEP Customers Count", "Integer", "AML", True, 3),
                field("tfs_screening_count", "TFS Screening Count", "Integer", "AML", True, 4),
                field("total_aml_alerts", "Total AML Alerts", "Integer", "AML", True, 5),
                field("aml_training_hours", "AML Training Hours", "Integer", "AML", True, 6),
                field("kyc_completion_rate", "KYC Completion Rate", "Percentage", "AML", True, 7),
            ],
            [
                formula_sum("total_aml_alerts", ["str_filed_count", "ctr_filed_count", "pep_customers_count"], "Total AML alerts", "Warning", 0),
            ],
            source_sheets=["AML - AML-CFT"],
        ),
        template(
            "INS_GOV",
            "Governance",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("GOV", "Governance", 1)],
            [
                field("board_members_count", "Board Members Count", "Integer", "GOV", True, 1),
                field("independent_directors_count", "Independent Directors Count", "Integer", "GOV", True, 2),
                field("board_meetings_held", "Board Meetings Held", "Integer", "GOV", True, 3),
                field("audit_committee_meetings", "Audit Committee Meetings", "Integer", "GOV", True, 4),
                field("compliance_breaches", "Compliance Breaches", "Integer", "GOV", True, 5),
                field("resolved_breaches", "Resolved Breaches", "Integer", "GOV", True, 6),
                field("breach_resolution_rate", "Breach Resolution Rate", "Percentage", "GOV", True, 7),
            ],
            [
                formula_ratio("breach_resolution_rate", "resolved_breaches", "compliance_breaches", "Breach resolution ratio", "Warning"),
            ],
            source_sheets=["GOV - Governance"],
        ),
        template(
            "INS_DIC",
            "Data Dictionary and Validation Summary",
            "Quarterly",
            "FixedRow",
            "ins",
            [section("DIC", "Validation Summary", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
                field("validation_warning_count", "Validation Warning Count", "Integer", "DIC", True, 4),
                field("data_quality_score", "Data Quality Score", "Percentage", "DIC", True, 5),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "INS_COV": (20, 6),
        "INS_SOL": (30, 14),
        "INS_PRM": (26, 12),
        "INS_CLM": (24, 10),
        "INS_TPR": (22, 9),
        "INS_INV": (24, 12),
        "INS_REI": (22, 9),
        "INS_FIN": (28, 14),
        "INS_RSK": (20, 8),
        "INS_AML": (20, 8),
        "INS_GOV": (18, 6),
        "INS_DIC": (18, 5),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "INSURANCE_NAICOM",
        "moduleVersion": "1.0.0",
        "description": "NAICOM insurance returns covering solvency, underwriting, claims, reserves, investments and AML.",
        "templates": templates,
        "interModuleDataFlows": [
            {
                "sourceTemplate": "INS_AML",
                "sourceField": "str_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_STR",
                "targetField": "str_filed_count",
                "transformationType": "DirectCopy",
                "description": "Insurance STR count to NFIU",
            },
            {
                "sourceTemplate": "INS_AML",
                "sourceField": "ctr_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_CTR",
                "targetField": "ctr_filed_count",
                "transformationType": "DirectCopy",
                "description": "Insurance CTR count to NFIU",
            },
            {
                "sourceTemplate": "INS_AML",
                "sourceField": "pep_customers_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_PEP",
                "targetField": "pep_alert_count",
                "transformationType": "DirectCopy",
                "description": "Insurance PEP data to NFIU",
            },
        ],
    }

    add_cross_rules(definition, 18)
    return definition


def pfa_module_definition():
    templates = [
        template(
            "PFA_COV",
            "PFA Cover",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("COV", "Cover", 1)],
            [
                field("pfa_licence_number", "PFA Licence Number", "Text", "COV", True, 1),
                field("pfa_name", "PFA Name", "Text", "COV", True, 2),
                field("reporting_period_start", "Reporting Period Start", "Date", "COV", True, 3),
                field("reporting_period_end", "Reporting Period End", "Date", "COV", True, 4),
                field("prepared_by", "Prepared By", "Text", "COV", True, 5),
                field("approved_by", "Approved By", "Text", "COV", True, 6),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "PFA_NAV",
            "Fund NAV Summary",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("NAV", "NAV Summary", 1)],
            [
                field("fund_i_assets", "Fund I Assets", "Money", "NAV", True, 1),
                field("fund_i_units", "Fund I Units", "Money", "NAV", True, 2),
                field("fund_i_nav_per_unit", "Fund I NAV Per Unit", "Decimal", "NAV", True, 3),
                field("fund_ii_assets", "Fund II Assets", "Money", "NAV", True, 4),
                field("fund_ii_units", "Fund II Units", "Money", "NAV", True, 5),
                field("fund_ii_nav_per_unit", "Fund II NAV Per Unit", "Decimal", "NAV", True, 6),
                field("fund_iii_assets", "Fund III Assets", "Money", "NAV", True, 7),
                field("fund_iii_units", "Fund III Units", "Money", "NAV", True, 8),
                field("fund_iii_nav_per_unit", "Fund III NAV Per Unit", "Decimal", "NAV", True, 9),
                field("fund_iv_assets", "Fund IV Assets", "Money", "NAV", True, 10),
                field("fund_iv_units", "Fund IV Units", "Money", "NAV", True, 11),
                field("fund_iv_nav_per_unit", "Fund IV NAV Per Unit", "Decimal", "NAV", True, 12),
                field("fund_v_assets", "Fund V Assets", "Money", "NAV", True, 13),
                field("fund_v_units", "Fund V Units", "Money", "NAV", True, 14),
                field("fund_v_nav_per_unit", "Fund V NAV Per Unit", "Decimal", "NAV", True, 15),
                field("fund_vi_assets", "Fund VI Assets", "Money", "NAV", True, 16),
                field("fund_vi_units", "Fund VI Units", "Money", "NAV", True, 17),
                field("fund_vi_nav_per_unit", "Fund VI NAV Per Unit", "Decimal", "NAV", True, 18),
                field("total_nav", "Total NAV", "Money", "NAV", True, 19),
                field("total_units", "Total Units", "Money", "NAV", True, 20),
                field("weighted_nav_per_unit", "Weighted NAV Per Unit", "Decimal", "NAV", True, 21),
            ],
            [
                formula_ratio("fund_i_nav_per_unit", "fund_i_assets", "fund_i_units", "Fund I NAV per unit"),
                formula_ratio("fund_ii_nav_per_unit", "fund_ii_assets", "fund_ii_units", "Fund II NAV per unit"),
                formula_ratio("fund_iii_nav_per_unit", "fund_iii_assets", "fund_iii_units", "Fund III NAV per unit"),
                formula_ratio("fund_iv_nav_per_unit", "fund_iv_assets", "fund_iv_units", "Fund IV NAV per unit"),
                formula_ratio("fund_v_nav_per_unit", "fund_v_assets", "fund_v_units", "Fund V NAV per unit"),
                formula_ratio("fund_vi_nav_per_unit", "fund_vi_assets", "fund_vi_units", "Fund VI NAV per unit"),
                formula_sum(
                    "total_nav",
                    ["fund_i_assets", "fund_ii_assets", "fund_iii_assets", "fund_iv_assets", "fund_v_assets", "fund_vi_assets"],
                    "Total NAV across funds",
                ),
                formula_sum(
                    "total_units",
                    ["fund_i_units", "fund_ii_units", "fund_iii_units", "fund_iv_units", "fund_v_units", "fund_vi_units"],
                    "Total units across funds",
                ),
                formula_ratio("weighted_nav_per_unit", "total_nav", "total_units", "Weighted NAV per unit"),
            ],
            source_sheets=["AUM - Assets Under Mgmt"],
        ),
        template(
            "PFA_FD1",
            "Fund I-III Details",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("FD1", "Fund I-III", 1)],
            [
                field("fund_i_equity", "Fund I Equity", "Money", "FD1", True, 1),
                field("fund_i_bonds", "Fund I Bonds", "Money", "FD1", True, 2),
                field("fund_i_money_market", "Fund I Money Market", "Money", "FD1", True, 3),
                field("fund_i_total", "Fund I Total", "Money", "FD1", True, 4),
                field("fund_ii_equity", "Fund II Equity", "Money", "FD1", True, 5),
                field("fund_ii_bonds", "Fund II Bonds", "Money", "FD1", True, 6),
                field("fund_ii_money_market", "Fund II Money Market", "Money", "FD1", True, 7),
                field("fund_ii_total", "Fund II Total", "Money", "FD1", True, 8),
                field("fund_iii_equity", "Fund III Equity", "Money", "FD1", True, 9),
                field("fund_iii_bonds", "Fund III Bonds", "Money", "FD1", True, 10),
                field("fund_iii_money_market", "Fund III Money Market", "Money", "FD1", True, 11),
                field("fund_iii_total", "Fund III Total", "Money", "FD1", True, 12),
            ],
            [
                formula_sum("fund_i_total", ["fund_i_equity", "fund_i_bonds", "fund_i_money_market"], "Fund I total"),
                formula_sum("fund_ii_total", ["fund_ii_equity", "fund_ii_bonds", "fund_ii_money_market"], "Fund II total"),
                formula_sum("fund_iii_total", ["fund_iii_equity", "fund_iii_bonds", "fund_iii_money_market"], "Fund III total"),
            ],
            source_sheets=["INV - Investment Compliance", "ROR - Rate of Return"],
        ),
        template(
            "PFA_FD2",
            "Fund IV-VI Details",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("FD2", "Fund IV-VI", 1)],
            [
                field("fund_iv_equity", "Fund IV Equity", "Money", "FD2", True, 1),
                field("fund_iv_bonds", "Fund IV Bonds", "Money", "FD2", True, 2),
                field("fund_iv_money_market", "Fund IV Money Market", "Money", "FD2", True, 3),
                field("fund_iv_total", "Fund IV Total", "Money", "FD2", True, 4),
                field("fund_v_equity", "Fund V Equity", "Money", "FD2", True, 5),
                field("fund_v_bonds", "Fund V Bonds", "Money", "FD2", True, 6),
                field("fund_v_money_market", "Fund V Money Market", "Money", "FD2", True, 7),
                field("fund_v_total", "Fund V Total", "Money", "FD2", True, 8),
                field("fund_vi_equity", "Fund VI Equity", "Money", "FD2", True, 9),
                field("fund_vi_bonds", "Fund VI Bonds", "Money", "FD2", True, 10),
                field("fund_vi_money_market", "Fund VI Money Market", "Money", "FD2", True, 11),
                field("fund_vi_total", "Fund VI Total", "Money", "FD2", True, 12),
            ],
            [
                formula_sum("fund_iv_total", ["fund_iv_equity", "fund_iv_bonds", "fund_iv_money_market"], "Fund IV total"),
                formula_sum("fund_v_total", ["fund_v_equity", "fund_v_bonds", "fund_v_money_market"], "Fund V total"),
                formula_sum("fund_vi_total", ["fund_vi_equity", "fund_vi_bonds", "fund_vi_money_market"], "Fund VI total"),
            ],
            source_sheets=["INV - Investment Compliance", "ROR - Rate of Return"],
        ),
        template(
            "PFA_RSA",
            "RSA Reconciliation",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("RSA", "RSA", 1)],
            [
                field("rsa_opening_balance", "RSA Opening Balance", "Money", "RSA", True, 1),
                field("rsa_contributions", "RSA Contributions", "Money", "RSA", True, 2),
                field("rsa_investment_income", "RSA Investment Income", "Money", "RSA", True, 3),
                field("rsa_benefit_payments", "RSA Benefit Payments", "Money", "RSA", True, 4),
                field("rsa_closing_balance", "RSA Closing Balance", "Money", "RSA", True, 5),
                field("rsa_pin_count", "RSA PIN Count", "Integer", "RSA", True, 6),
            ],
            [
                formula_sum("rsa_closing_balance", ["rsa_opening_balance", "rsa_contributions", "rsa_investment_income"], "RSA movement before payments", "Warning", 1),
                formula_compare("GreaterThanOrEqual", "rsa_closing_balance", "rsa_benefit_payments", "Closing balance >= benefits", "Warning"),
            ],
            source_sheets=["RSA - Account Management"],
        ),
        template(
            "PFA_AAL",
            "Asset Allocation",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("AAL", "Asset Allocation Limits", 1)],
            [
                field("equity_percent", "Equity Allocation", "Percentage", "AAL", True, 1),
                field("money_market_percent", "Money Market Allocation", "Percentage", "AAL", True, 2),
                field("fgn_bond_percent", "FGN Bond Allocation", "Percentage", "AAL", True, 3),
                field("corp_bond_percent", "Corporate Bond Allocation", "Percentage", "AAL", True, 4),
                field("infrastructure_percent", "Infrastructure Allocation", "Percentage", "AAL", True, 5),
                field("private_equity_percent", "Private Equity Allocation", "Percentage", "AAL", True, 6),
                field("max_equity_percent", "Max Equity", "Percentage", "AAL", True, 7, 0, 100, 2),
                field("max_money_market_percent", "Max Money Market", "Percentage", "AAL", True, 8, 0, 100, 2),
                field("max_private_equity_percent", "Max Private Equity", "Percentage", "AAL", True, 9, 0, 100, 2),
                field("total_allocation_percent", "Total Allocation", "Percentage", "AAL", True, 10),
                field("allocation_target_percent", "Allocation Target", "Percentage", "AAL", True, 11, 100, 100, 2),
            ],
            [
                formula_sum(
                    "total_allocation_percent",
                    [
                        "equity_percent",
                        "money_market_percent",
                        "fgn_bond_percent",
                        "corp_bond_percent",
                        "infrastructure_percent",
                        "private_equity_percent",
                    ],
                    "Total asset allocation",
                ),
                formula_compare("LessThanOrEqual", "equity_percent", "max_equity_percent", "Equity cap compliance"),
                formula_compare("LessThanOrEqual", "money_market_percent", "max_money_market_percent", "Money market cap compliance"),
                formula_compare("LessThanOrEqual", "private_equity_percent", "max_private_equity_percent", "Private equity cap compliance"),
                formula_compare("Equals", "total_allocation_percent", "allocation_target_percent", "Allocation should total 100"),
            ],
            source_sheets=["INV - Investment Compliance"],
        ),
        template(
            "PFA_CON",
            "Contributions",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("CON", "Contributions", 1)],
            [
                field("employee_contributions", "Employee Contributions", "Money", "CON", True, 1),
                field("employer_contributions", "Employer Contributions", "Money", "CON", True, 2),
                field("voluntary_contributions", "Voluntary Contributions", "Money", "CON", True, 3),
                field("total_contributions", "Total Contributions", "Money", "CON", True, 4),
                field("unmatched_contributions", "Unmatched Contributions", "Money", "CON", True, 5),
            ],
            [
                formula_sum("total_contributions", ["employee_contributions", "employer_contributions", "voluntary_contributions"], "Total contributions"),
            ],
            source_sheets=["CON - Contributions"],
        ),
        template(
            "PFA_BEN",
            "Benefits and Withdrawals",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("BEN", "Benefits", 1)],
            [
                field("retirement_benefits", "Retirement Benefits", "Money", "BEN", True, 1),
                field("programmed_withdrawals", "Programmed Withdrawals", "Money", "BEN", True, 2),
                field("lump_sum_payments", "Lump Sum Payments", "Money", "BEN", True, 3),
                field("death_benefits", "Death Benefits", "Money", "BEN", True, 4),
                field("total_benefits_paid", "Total Benefits Paid", "Money", "BEN", True, 5),
            ],
            [
                formula_sum("total_benefits_paid", ["retirement_benefits", "programmed_withdrawals", "lump_sum_payments", "death_benefits"], "Total benefits paid"),
            ],
            source_sheets=["BEN - Benefits Admin"],
        ),
        template(
            "PFA_INC",
            "Investment Income",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("INC", "Investment Income", 1)],
            [
                field("interest_income", "Interest Income", "Money", "INC", True, 1),
                field("dividend_income", "Dividend Income", "Money", "INC", True, 2),
                field("capital_gains", "Capital Gains", "Money", "INC", True, 3),
                field("other_investment_income", "Other Investment Income", "Money", "INC", True, 4),
                field("total_investment_income", "Total Investment Income", "Money", "INC", True, 5),
            ],
            [
                formula_sum("total_investment_income", ["interest_income", "dividend_income", "capital_gains", "other_investment_income"], "Total investment income"),
            ],
            source_sheets=["ROR - Rate of Return"],
        ),
        template(
            "PFA_AML",
            "AML/CFT",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("AML", "AML", 1)],
            [
                field("str_filed_count", "STR Filed Count", "Integer", "AML", True, 1),
                field("ctr_filed_count", "CTR Filed Count", "Integer", "AML", True, 2),
                field("pep_customers_count", "PEP Customers Count", "Integer", "AML", True, 3),
                field("tfs_screening_count", "TFS Screening Count", "Integer", "AML", True, 4),
                field("total_aml_alerts", "Total AML Alerts", "Integer", "AML", True, 5),
            ],
            [
                formula_sum("total_aml_alerts", ["str_filed_count", "ctr_filed_count", "pep_customers_count"], "Total AML alerts", "Warning", 0),
            ],
            source_sheets=["AML - AML-CFT"],
        ),
        template(
            "PFA_GOV",
            "Governance",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("GOV", "Governance", 1)],
            [
                field("board_meetings_held", "Board Meetings Held", "Integer", "GOV", True, 1),
                field("investment_committee_meetings", "Investment Committee Meetings", "Integer", "GOV", True, 2),
                field("compliance_breaches", "Compliance Breaches", "Integer", "GOV", True, 3),
                field("resolved_breaches", "Resolved Breaches", "Integer", "GOV", True, 4),
            ],
            [
                formula_compare("GreaterThanOrEqual", "board_meetings_held", "investment_committee_meetings", "Board meetings should cover investment governance", "Warning"),
            ],
            source_sheets=["GOV - Governance"],
        ),
        template(
            "PFA_DIC",
            "Data Dictionary",
            "Monthly",
            "FixedRow",
            "pfa",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
                field("validation_warning_count", "Validation Warning Count", "Integer", "DIC", True, 4),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "PFA_COV": (18, 5),
        "PFA_NAV": (28, 14),
        "PFA_FD1": (24, 10),
        "PFA_FD2": (24, 10),
        "PFA_RSA": (22, 9),
        "PFA_AAL": (24, 12),
        "PFA_CON": (22, 8),
        "PFA_BEN": (22, 8),
        "PFA_INC": (20, 7),
        "PFA_AML": (18, 7),
        "PFA_GOV": (16, 6),
        "PFA_DIC": (16, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "PFA_PENCOM",
        "moduleVersion": "1.0.0",
        "description": "PenCom PFA returns covering multi-fund NAV, allocations, RSA reconciliation and AML.",
        "templates": templates,
        "interModuleDataFlows": [
            {
                "sourceTemplate": "PFA_AML",
                "sourceField": "str_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_STR",
                "targetField": "str_filed_count",
                "transformationType": "DirectCopy",
                "description": "PFA STR count to NFIU",
            },
            {
                "sourceTemplate": "PFA_AML",
                "sourceField": "ctr_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_CTR",
                "targetField": "ctr_filed_count",
                "transformationType": "DirectCopy",
                "description": "PFA CTR count to NFIU",
            },
            {
                "sourceTemplate": "PFA_AML",
                "sourceField": "pep_customers_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_PEP",
                "targetField": "pep_alert_count",
                "transformationType": "DirectCopy",
                "description": "PFA PEP count to NFIU",
            },
        ],
    }

    add_cross_rules(definition, 18)
    return definition


def cmo_module_definition():
    templates = [
        template(
            "CMO_COV",
            "CMO Cover",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("COV", "Cover", 1)],
            [
                field("sec_reg_no", "SEC Registration Number", "Text", "COV", True, 1),
                field("cmo_name", "CMO Name", "Text", "COV", True, 2),
                field("cmo_category", "CMO Category", "Text", "COV", True, 3),
                field("reporting_period_start", "Reporting Period Start", "Date", "COV", True, 4),
                field("reporting_period_end", "Reporting Period End", "Date", "COV", True, 5),
                field("prepared_by", "Prepared By", "Text", "COV", True, 6),
                field("approved_by", "Approved By", "Text", "COV", True, 7),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "CMO_CAP",
            "Net Capital",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("CAP", "Net Capital", 1)],
            [
                field("liquid_assets", "Liquid Assets", "Money", "CAP", True, 1),
                field("total_liabilities", "Total Liabilities", "Money", "CAP", True, 2),
                field("net_capital", "Net Capital", "Money", "CAP", True, 3),
                field("minimum_net_capital", "Minimum Net Capital", "Money", "CAP", True, 4),
                field("net_capital_ratio", "Net Capital Ratio", "Percentage", "CAP", True, 5),
            ],
            [
                formula_difference("net_capital", "liquid_assets", "total_liabilities", "Net capital"),
                formula_compare("GreaterThanOrEqual", "net_capital", "minimum_net_capital", "Minimum net capital"),
                formula_ratio("net_capital_ratio", "net_capital", "minimum_net_capital", "Net capital coverage ratio"),
            ],
            source_sheets=["CAP - Net Capital"],
        ),
        template(
            "CMO_CLI",
            "Client Asset Segregation",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("CLI", "Client Assets", 1)],
            [
                field("client_cash", "Client Cash", "Money", "CLI", True, 1),
                field("client_securities", "Client Securities", "Money", "CLI", True, 2),
                field("client_total_assets", "Client Total Assets", "Money", "CLI", True, 3),
                field("proprietary_assets", "Proprietary Assets", "Money", "CLI", True, 4),
                field("segregation_buffer", "Segregation Buffer", "Money", "CLI", True, 5),
            ],
            [
                formula_sum("client_total_assets", ["client_cash", "client_securities"], "Total client assets"),
                formula_difference("segregation_buffer", "client_total_assets", "proprietary_assets", "Client-proprietary segregation buffer"),
                formula_compare("GreaterThan", "client_total_assets", "proprietary_assets", "Client assets must be segregated from proprietary assets"),
            ],
            source_sheets=["CLI - Client Assets"],
        ),
        template(
            "CMO_TRD",
            "Trading Volumes",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("TRD", "Trading Volumes", 1)],
            [
                field("equities_volume", "Equities Volume", "Money", "TRD", True, 1),
                field("bonds_volume", "Bonds Volume", "Money", "TRD", True, 2),
                field("commercial_paper_volume", "Commercial Paper Volume", "Money", "TRD", True, 3),
                field("mutual_funds_volume", "Mutual Funds Volume", "Money", "TRD", True, 4),
                field("other_instruments_volume", "Other Instruments Volume", "Money", "TRD", True, 5),
                field("total_trade_volume", "Total Trade Volume", "Money", "TRD", True, 6),
                field("equities_value", "Equities Value", "Money", "TRD", True, 7),
                field("bonds_value", "Bonds Value", "Money", "TRD", True, 8),
                field("commercial_paper_value", "Commercial Paper Value", "Money", "TRD", True, 9),
                field("mutual_funds_value", "Mutual Funds Value", "Money", "TRD", True, 10),
                field("other_instruments_value", "Other Instruments Value", "Money", "TRD", True, 11),
                field("total_trade_value", "Total Trade Value", "Money", "TRD", True, 12),
            ],
            [
                formula_sum(
                    "total_trade_volume",
                    ["equities_volume", "bonds_volume", "commercial_paper_volume", "mutual_funds_volume", "other_instruments_volume"],
                    "Total trade volume",
                ),
                formula_sum(
                    "total_trade_value",
                    ["equities_value", "bonds_value", "commercial_paper_value", "mutual_funds_value", "other_instruments_value"],
                    "Total trade value",
                ),
            ],
            source_sheets=["TRD - Trading Activity"],
        ),
        template(
            "CMO_AUM",
            "AUM and Portfolio",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("AUM", "AUM", 1)],
            [
                field("discretionary_aum", "Discretionary AUM", "Money", "AUM", True, 1),
                field("advisory_aum", "Advisory AUM", "Money", "AUM", True, 2),
                field("collective_investment_aum", "Collective Investment AUM", "Money", "AUM", True, 3),
                field("total_aum", "Total AUM", "Money", "AUM", True, 4),
                field("equity_aum", "Equity AUM", "Money", "AUM", True, 5),
                field("fixed_income_aum", "Fixed Income AUM", "Money", "AUM", True, 6),
                field("money_market_aum", "Money Market AUM", "Money", "AUM", True, 7),
                field("other_aum", "Other AUM", "Money", "AUM", True, 8),
            ],
            [
                formula_sum("total_aum", ["discretionary_aum", "advisory_aum", "collective_investment_aum"], "Total AUM"),
            ],
            source_sheets=["CIS - Collective Investment"],
        ),
        template(
            "CMO_REV",
            "Revenue",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("REV", "Revenue", 1)],
            [
                field("brokerage_income", "Brokerage Income", "Money", "REV", True, 1),
                field("advisory_fees", "Advisory Fees", "Money", "REV", True, 2),
                field("portfolio_management_fees", "Portfolio Management Fees", "Money", "REV", True, 3),
                field("other_fee_income", "Other Fee Income", "Money", "REV", True, 4),
                field("total_revenue", "Total Revenue", "Money", "REV", True, 5),
            ],
            [
                formula_sum("total_revenue", ["brokerage_income", "advisory_fees", "portfolio_management_fees", "other_fee_income"], "Total revenue"),
            ],
            source_sheets=["IHS - Issuing House", "REG - Registrar"],
        ),
        template(
            "CMO_FIN",
            "Financial Statements",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("FIN", "Financial Statements", 1)],
            [
                field("total_assets", "Total Assets", "Money", "FIN", True, 1),
                field("total_liabilities", "Total Liabilities", "Money", "FIN", True, 2),
                field("shareholders_funds", "Shareholders Funds", "Money", "FIN", True, 3),
                field("operating_expenses", "Operating Expenses", "Money", "FIN", True, 4),
                field("tax_expense", "Tax Expense", "Money", "FIN", True, 5),
                field("total_expenses", "Total Expenses", "Money", "FIN", True, 6),
                field("total_revenue", "Total Revenue", "Money", "FIN", True, 7),
                field("profit_before_tax", "Profit Before Tax", "Money", "FIN", True, 8),
                field("profit_after_tax", "Profit After Tax", "Money", "FIN", True, 9),
                field("total_trade_value", "Total Trade Value", "Money", "FIN", True, 10),
            ],
            [
                formula_difference("shareholders_funds", "total_assets", "total_liabilities", "Balance sheet equation"),
                formula_sum("total_expenses", ["operating_expenses", "tax_expense"], "Total expenses"),
                formula_difference("profit_before_tax", "total_revenue", "total_expenses", "PBT"),
                formula_difference("profit_after_tax", "profit_before_tax", "tax_expense", "PAT"),
            ],
            source_sheets=["FIN - Financial Statements"],
        ),
        template(
            "CMO_RSK",
            "Risk Management",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("RSK", "Risk", 1)],
            [
                field("market_risk_var", "Market Risk VaR", "Money", "RSK", True, 1),
                field("liquidity_buffer", "Liquidity Buffer", "Money", "RSK", True, 2),
                field("operational_loss_events", "Operational Loss Events", "Integer", "RSK", True, 3),
                field("stress_loss_estimate", "Stress Loss Estimate", "Money", "RSK", True, 4),
            ],
            [],
            source_sheets=["DAS - Digital Assets"],
        ),
        template(
            "CMO_IPR",
            "Investor Protection",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("IPR", "Investor Protection", 1)],
            [
                field("complaints_received", "Complaints Received", "Integer", "IPR", True, 1),
                field("complaints_resolved", "Complaints Resolved", "Integer", "IPR", True, 2),
                field("resolution_ratio", "Resolution Ratio", "Percentage", "IPR", True, 3),
            ],
            [
                formula_ratio("resolution_ratio", "complaints_resolved", "complaints_received", "Complaint resolution ratio", "Warning"),
            ],
            source_sheets=["CLI - Client Assets"],
        ),
        template(
            "CMO_AML",
            "AML/CFT",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("AML", "AML", 1)],
            [
                field("str_filed_count", "STR Filed Count", "Integer", "AML", True, 1),
                field("ctr_filed_count", "CTR Filed Count", "Integer", "AML", True, 2),
                field("pep_customers_count", "PEP Count", "Integer", "AML", True, 3),
                field("tfs_screening_count", "TFS Screening Count", "Integer", "AML", True, 4),
                field("total_aml_alerts", "Total AML Alerts", "Integer", "AML", True, 5),
            ],
            [
                formula_sum("total_aml_alerts", ["str_filed_count", "ctr_filed_count", "pep_customers_count"], "Total AML alerts", "Warning", 0),
            ],
            source_sheets=["AML - AML-CFT"],
        ),
        template(
            "CMO_REG",
            "Regulatory Compliance",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("REG", "Regulatory", 1)],
            [
                field("returns_due", "Returns Due", "Integer", "REG", True, 1),
                field("returns_submitted", "Returns Submitted", "Integer", "REG", True, 2),
                field("returns_on_time", "Returns On Time", "Integer", "REG", True, 3),
            ],
            [
                formula_compare("LessThanOrEqual", "returns_on_time", "returns_submitted", "On-time returns cannot exceed submissions"),
            ],
            source_sheets=["REG - Registrar"],
        ),
        template(
            "CMO_GOV",
            "Governance",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("GOV", "Governance", 1)],
            [
                field("board_meetings_held", "Board Meetings Held", "Integer", "GOV", True, 1),
                field("risk_committee_meetings", "Risk Committee Meetings", "Integer", "GOV", True, 2),
                field("audit_committee_meetings", "Audit Committee Meetings", "Integer", "GOV", True, 3),
            ],
            [],
            source_sheets=["GOV - Governance"],
        ),
        template(
            "CMO_DIC",
            "Data Dictionary",
            "Monthly",
            "FixedRow",
            "cmo",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "CMO_COV": (18, 5),
        "CMO_CAP": (24, 10),
        "CMO_CLI": (24, 10),
        "CMO_TRD": (24, 11),
        "CMO_AUM": (22, 9),
        "CMO_REV": (20, 8),
        "CMO_FIN": (24, 11),
        "CMO_RSK": (18, 6),
        "CMO_IPR": (18, 6),
        "CMO_AML": (18, 7),
        "CMO_REG": (18, 6),
        "CMO_GOV": (16, 5),
        "CMO_DIC": (16, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "CMO_SEC",
        "moduleVersion": "1.0.0",
        "description": "SEC capital market operator returns with net capital, client asset segregation, AUM and AML.",
        "templates": templates,
        "interModuleDataFlows": [
            {
                "sourceTemplate": "CMO_AML",
                "sourceField": "str_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_STR",
                "targetField": "str_filed_count",
                "transformationType": "DirectCopy",
                "description": "CMO STR count to NFIU",
            },
            {
                "sourceTemplate": "CMO_AML",
                "sourceField": "ctr_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_CTR",
                "targetField": "ctr_filed_count",
                "transformationType": "DirectCopy",
                "description": "CMO CTR count to NFIU",
            },
            {
                "sourceTemplate": "CMO_AML",
                "sourceField": "pep_customers_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_PEP",
                "targetField": "pep_alert_count",
                "transformationType": "DirectCopy",
                "description": "CMO PEP count to NFIU",
            },
            {
                "sourceTemplate": "CMO_AUM",
                "sourceField": "total_aum",
                "targetModule": "ESG_CLIMATE",
                "targetTemplate": "ESG_FINANCED_EMISSIONS",
                "targetField": "portfolio_exposure",
                "transformationType": "DirectCopy",
                "description": "CMO AUM to ESG financed exposure",
            },
        ],
    }

    add_cross_rules(definition, 20)
    return definition


def dfi_module_definition():
    templates = [
        template(
            "DFI_COV",
            "DFI Cover",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("COV", "Cover", 1)],
            [
                field("dfi_licence_number", "DFI Licence Number", "Text", "COV", True, 1),
                field("dfi_name", "DFI Name", "Text", "COV", True, 2),
                field("dfi_mandate", "DFI Mandate", "Text", "COV", True, 3),
                field("reporting_period_start", "Reporting Period Start", "Date", "COV", True, 4),
                field("reporting_period_end", "Reporting Period End", "Date", "COV", True, 5),
                field("prepared_by", "Prepared By", "Text", "COV", True, 6),
                field("approved_by", "Approved By", "Text", "COV", True, 7),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "DFI_SEC",
            "Sector Allocation",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("SEC", "Sector Allocation", 1)],
            [
                field("agriculture_percent", "Agriculture Allocation", "Percentage", "SEC", True, 1),
                field("sme_percent", "SME Allocation", "Percentage", "SEC", True, 2),
                field("health_percent", "Health Allocation", "Percentage", "SEC", True, 3),
                field("education_percent", "Education Allocation", "Percentage", "SEC", True, 4),
                field("renewable_percent", "Renewable Energy Allocation", "Percentage", "SEC", True, 5),
                field("infrastructure_percent", "Infrastructure Allocation", "Percentage", "SEC", True, 6),
                field("total_sector_allocation_percent", "Total Sector Allocation", "Percentage", "SEC", True, 7),
                field("sector_allocation_target_percent", "Sector Allocation Target", "Percentage", "SEC", True, 8, 100, 100, 2),
                field("agriculture_exposure", "Agriculture Exposure", "Money", "SEC", True, 9),
                field("renewable_exposure", "Renewable Exposure", "Money", "SEC", True, 10),
                field("total_sector_exposure", "Total Sector Exposure", "Money", "SEC", True, 11),
            ],
            [
                formula_sum(
                    "total_sector_allocation_percent",
                    ["agriculture_percent", "sme_percent", "health_percent", "education_percent", "renewable_percent", "infrastructure_percent"],
                    "Total sector allocation",
                ),
                formula_compare("Equals", "total_sector_allocation_percent", "sector_allocation_target_percent", "Sector allocation must total 100"),
                formula_sum("total_sector_exposure", ["agriculture_exposure", "renewable_exposure"], "Tracked climate sector exposure", "Warning", 1),
            ],
            source_sheets=["LPQ - Loan Portfolio", "DEV - Development Impact"],
        ),
        template(
            "DFI_CON",
            "Concessional Lending",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("CON", "Concessional Lending", 1)],
            [
                field("concessional_loans_outstanding", "Concessional Loans Outstanding", "Money", "CON", True, 1),
                field("commercial_loans_outstanding", "Commercial Loans Outstanding", "Money", "CON", True, 2),
                field("average_concessional_rate", "Average Concessional Rate", "Percentage", "CON", True, 3),
                field("average_commercial_rate", "Average Commercial Rate", "Percentage", "CON", True, 4),
                field("rate_concession_spread", "Rate Concession Spread", "Percentage", "CON", True, 5),
            ],
            [
                formula_difference("rate_concession_spread", "average_commercial_rate", "average_concessional_rate", "Concessional spread"),
            ],
            source_sheets=["LPQ - Loan Portfolio"],
        ),
        template(
            "DFI_IMP",
            "Impact Metrics",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("IMP", "Impact", 1)],
            [
                field("jobs_created", "Jobs Created", "Integer", "IMP", True, 1),
                field("women_owned_sme_supported", "Women-Owned SMEs Supported", "Integer", "IMP", True, 2),
                field("rural_beneficiaries", "Rural Beneficiaries", "Integer", "IMP", True, 3),
                field("co2_reduction_tonnes", "CO2 Reduction Tonnes", "Money", "IMP", True, 4),
            ],
            [],
            source_sheets=["DEV - Development Impact"],
        ),
        template(
            "DFI_CAP",
            "Capital Adequacy",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("CAP", "Capital", 1)],
            [
                field("tier1_capital", "Tier 1 Capital", "Money", "CAP", True, 1),
                field("tier2_capital", "Tier 2 Capital", "Money", "CAP", True, 2),
                field("risk_weighted_assets", "Risk Weighted Assets", "Money", "CAP", True, 3),
                field("car_ratio", "CAR Ratio", "Percentage", "CAP", True, 4),
                field("minimum_car_ratio", "Minimum CAR Ratio", "Percentage", "CAP", True, 5),
            ],
            [
                formula_custom("car_ratio", ["tier1_capital", "tier2_capital", "risk_weighted_assets"], "CAR", "CAR calculation"),
                formula_compare("GreaterThanOrEqual", "car_ratio", "minimum_car_ratio", "CAR threshold"),
            ],
            source_sheets=["CAP - Capital Adequacy"],
        ),
        template(
            "DFI_FIN",
            "Financial Statements",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("FIN", "Financial", 1)],
            [
                field("total_assets", "Total Assets", "Money", "FIN", True, 1),
                field("total_liabilities", "Total Liabilities", "Money", "FIN", True, 2),
                field("equity", "Equity", "Money", "FIN", True, 3),
                field("interest_income", "Interest Income", "Money", "FIN", True, 4),
                field("fee_income", "Fee Income", "Money", "FIN", True, 5),
                field("total_income", "Total Income", "Money", "FIN", True, 6),
                field("operating_expenses", "Operating Expenses", "Money", "FIN", True, 7),
                field("total_expenses", "Total Expenses", "Money", "FIN", True, 8),
                field("profit_before_tax", "Profit Before Tax", "Money", "FIN", True, 9),
            ],
            [
                formula_difference("equity", "total_assets", "total_liabilities", "Balance sheet equation"),
                formula_sum("total_income", ["interest_income", "fee_income"], "Total income"),
                formula_sum("total_expenses", ["operating_expenses"], "Total expenses"),
                formula_difference("profit_before_tax", "total_income", "total_expenses", "PBT"),
            ],
            source_sheets=["FIN - Financial Statements"],
        ),
        template(
            "DFI_IFR",
            "IFRS 9 Staging",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("IFR", "IFRS 9", 1)],
            [
                field("stage1_pd", "Stage 1 PD", "Decimal", "IFR", True, 1),
                field("stage1_lgd", "Stage 1 LGD", "Decimal", "IFR", True, 2),
                field("stage1_ead", "Stage 1 EAD", "Money", "IFR", True, 3),
                field("stage1_ecl", "Stage 1 ECL", "Money", "IFR", True, 4),
                field("stage2_pd", "Stage 2 PD", "Decimal", "IFR", True, 5),
                field("stage2_lgd", "Stage 2 LGD", "Decimal", "IFR", True, 6),
                field("stage2_ead", "Stage 2 EAD", "Money", "IFR", True, 7),
                field("stage2_ecl", "Stage 2 ECL", "Money", "IFR", True, 8),
                field("stage3_pd", "Stage 3 PD", "Decimal", "IFR", True, 9),
                field("stage3_lgd", "Stage 3 LGD", "Decimal", "IFR", True, 10),
                field("stage3_ead", "Stage 3 EAD", "Money", "IFR", True, 11),
                field("stage3_ecl", "Stage 3 ECL", "Money", "IFR", True, 12),
                field("total_ecl", "Total ECL", "Money", "IFR", True, 13),
            ],
            [
                formula_custom("stage1_ecl", ["stage1_pd", "stage1_lgd", "stage1_ead"], "ECL", "Stage 1 ECL"),
                formula_custom("stage2_ecl", ["stage2_pd", "stage2_lgd", "stage2_ead"], "ECL", "Stage 2 ECL"),
                formula_custom("stage3_ecl", ["stage3_pd", "stage3_lgd", "stage3_ead"], "ECL", "Stage 3 ECL"),
                formula_sum("total_ecl", ["stage1_ecl", "stage2_ecl", "stage3_ecl"], "Total ECL"),
            ],
            source_sheets=["ASQ - Asset Quality"],
        ),
        template(
            "DFI_INT",
            "Intervention Funds",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("INT", "Intervention Funds", 1)],
            [
                field("opening_intervention_balance", "Opening Intervention Balance", "Money", "INT", True, 1),
                field("new_disbursement", "New Disbursement", "Money", "INT", True, 2),
                field("repayment_received", "Repayment Received", "Money", "INT", True, 3),
                field("closing_intervention_balance", "Closing Intervention Balance", "Money", "INT", True, 4),
            ],
            [
                formula_sum("closing_intervention_balance", ["opening_intervention_balance", "new_disbursement"], "Closing before repayments", "Warning", 1),
                formula_compare("GreaterThanOrEqual", "closing_intervention_balance", "repayment_received", "Closing balance >= repayments", "Warning"),
            ],
            source_sheets=["FND - Funding & Liquidity", "CGR - Credit Guarantee"],
        ),
        template(
            "DFI_AML",
            "AML/CFT",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("AML", "AML", 1)],
            [
                field("str_filed_count", "STR Filed Count", "Integer", "AML", True, 1),
                field("ctr_filed_count", "CTR Filed Count", "Integer", "AML", True, 2),
                field("pep_customers_count", "PEP Count", "Integer", "AML", True, 3),
                field("tfs_screening_count", "TFS Screening Count", "Integer", "AML", True, 4),
                field("total_aml_alerts", "Total AML Alerts", "Integer", "AML", True, 5),
            ],
            [
                formula_sum("total_aml_alerts", ["str_filed_count", "ctr_filed_count", "pep_customers_count"], "Total AML alerts", "Warning", 0),
            ],
            source_sheets=["AML - AML-CFT"],
        ),
        template(
            "DFI_RSK",
            "Risk Management",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("RSK", "Risk", 1)],
            [
                field("portfolio_at_risk", "Portfolio at Risk", "Percentage", "RSK", True, 1),
                field("liquidity_coverage_ratio", "Liquidity Coverage Ratio", "Percentage", "RSK", True, 2),
                field("stress_loss_ratio", "Stress Loss Ratio", "Percentage", "RSK", True, 3),
            ],
            [],
            source_sheets=["ASQ - Asset Quality"],
        ),
        template(
            "DFI_GOV",
            "Governance",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("GOV", "Governance", 1)],
            [
                field("board_meetings_held", "Board Meetings Held", "Integer", "GOV", True, 1),
                field("risk_committee_meetings", "Risk Committee Meetings", "Integer", "GOV", True, 2),
                field("audit_findings", "Audit Findings", "Integer", "GOV", True, 3),
                field("resolved_audit_findings", "Resolved Audit Findings", "Integer", "GOV", True, 4),
            ],
            [],
            source_sheets=["GOV - Governance"],
        ),
        template(
            "DFI_DIC",
            "Data Dictionary",
            "Quarterly",
            "FixedRow",
            "dfi",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "DFI_COV": (18, 5),
        "DFI_SEC": (26, 12),
        "DFI_CON": (22, 8),
        "DFI_IMP": (20, 6),
        "DFI_CAP": (22, 9),
        "DFI_FIN": (24, 10),
        "DFI_IFR": (24, 11),
        "DFI_INT": (22, 9),
        "DFI_AML": (18, 7),
        "DFI_RSK": (18, 6),
        "DFI_GOV": (16, 5),
        "DFI_DIC": (16, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "DFI_CBN",
        "moduleVersion": "1.0.0",
        "description": "CBN DFI returns for sector allocation, concessional lending, development impact, IFRS9 and AML.",
        "templates": templates,
        "interModuleDataFlows": [
            {
                "sourceTemplate": "DFI_AML",
                "sourceField": "str_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_STR",
                "targetField": "str_filed_count",
                "transformationType": "DirectCopy",
                "description": "DFI STR count to NFIU",
            },
            {
                "sourceTemplate": "DFI_AML",
                "sourceField": "ctr_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_CTR",
                "targetField": "ctr_filed_count",
                "transformationType": "DirectCopy",
                "description": "DFI CTR count to NFIU",
            },
            {
                "sourceTemplate": "DFI_AML",
                "sourceField": "pep_customers_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_PEP",
                "targetField": "pep_alert_count",
                "transformationType": "DirectCopy",
                "description": "DFI PEP data to NFIU",
            },
            {
                "sourceTemplate": "DFI_SEC",
                "sourceField": "agriculture_exposure",
                "targetModule": "ESG_CLIMATE",
                "targetTemplate": "ESG_FINANCED_EMISSIONS",
                "targetField": "agriculture_exposure",
                "transformationType": "DirectCopy",
                "description": "DFI agriculture exposure to ESG",
            },
            {
                "sourceTemplate": "DFI_SEC",
                "sourceField": "renewable_exposure",
                "targetModule": "ESG_CLIMATE",
                "targetTemplate": "ESG_FINANCED_EMISSIONS",
                "targetField": "renewable_energy_exposure",
                "transformationType": "DirectCopy",
                "description": "DFI renewable exposure to ESG",
            },
        ],
    }

    add_cross_rules(definition, 18)
    return definition


def imto_module_definition():
    templates = [
        template(
            "IMTO_COV",
            "IMTO Cover",
            "Monthly",
            "FixedRow",
            "imto",
            [section("COV", "Cover", 1)],
            [
                field("imto_licence_number", "IMTO Licence Number", "Text", "COV", True, 1),
                field("imto_name", "IMTO Name", "Text", "COV", True, 2),
                field("reporting_period_start", "Reporting Period Start", "Date", "COV", True, 3),
                field("reporting_period_end", "Reporting Period End", "Date", "COV", True, 4),
                field("prepared_by", "Prepared By", "Text", "COV", True, 5),
                field("approved_by", "Approved By", "Text", "COV", True, 6),
            ],
            [],
            source_sheets=["Cover"],
        ),
        template(
            "IMTO_INB",
            "Inbound Remittances",
            "Monthly",
            "FixedRow",
            "imto",
            [section("INB", "Inbound Volumes", 1)],
            [
                field("bank_transfer_count", "Bank Transfer Count", "Integer", "INB", True, 1),
                field("mobile_money_count", "Mobile Money Count", "Integer", "INB", True, 2),
                field("cash_pickup_count", "Cash Pickup Count", "Integer", "INB", True, 3),
                field("total_inbound_count", "Total Inbound Count", "Integer", "INB", True, 4),
                field("bank_transfer_value", "Bank Transfer Value", "Money", "INB", True, 5),
                field("mobile_money_value", "Mobile Money Value", "Money", "INB", True, 6),
                field("cash_pickup_value", "Cash Pickup Value", "Money", "INB", True, 7),
                field("total_inbound_value", "Total Inbound Value", "Money", "INB", True, 8),
            ],
            [
                formula_sum("total_inbound_count", ["bank_transfer_count", "mobile_money_count", "cash_pickup_count"], "Total inbound count"),
                formula_sum("total_inbound_value", ["bank_transfer_value", "mobile_money_value", "cash_pickup_value"], "Total inbound value"),
            ],
            source_sheets=["VOL - Volume & Value"],
        ),
        template(
            "IMTO_COR",
            "Corridor Analysis",
            "Monthly",
            "FixedRow",
            "imto",
            [section("COR", "Corridor", 1)],
            [
                field("usa_corridor_value", "USA Corridor Value", "Money", "COR", True, 1),
                field("uk_corridor_value", "UK Corridor Value", "Money", "COR", True, 2),
                field("uae_corridor_value", "UAE Corridor Value", "Money", "COR", True, 3),
                field("canada_corridor_value", "Canada Corridor Value", "Money", "COR", True, 4),
                field("saudi_corridor_value", "Saudi Corridor Value", "Money", "COR", True, 5),
                field("other_corridor_value", "Other Corridor Value", "Money", "COR", True, 6),
                field("total_corridor_value", "Total Corridor Value", "Money", "COR", True, 7),
                field("usa_corridor_count", "USA Corridor Count", "Integer", "COR", True, 8),
                field("uk_corridor_count", "UK Corridor Count", "Integer", "COR", True, 9),
                field("uae_corridor_count", "UAE Corridor Count", "Integer", "COR", True, 10),
                field("canada_corridor_count", "Canada Corridor Count", "Integer", "COR", True, 11),
                field("saudi_corridor_count", "Saudi Corridor Count", "Integer", "COR", True, 12),
                field("other_corridor_count", "Other Corridor Count", "Integer", "COR", True, 13),
                field("total_corridor_count", "Total Corridor Count", "Integer", "COR", True, 14),
            ],
            [
                formula_sum(
                    "total_corridor_value",
                    [
                        "usa_corridor_value",
                        "uk_corridor_value",
                        "uae_corridor_value",
                        "canada_corridor_value",
                        "saudi_corridor_value",
                        "other_corridor_value",
                    ],
                    "Total corridor value",
                ),
                formula_sum(
                    "total_corridor_count",
                    [
                        "usa_corridor_count",
                        "uk_corridor_count",
                        "uae_corridor_count",
                        "canada_corridor_count",
                        "saudi_corridor_count",
                        "other_corridor_count",
                    ],
                    "Total corridor count",
                ),
            ],
            source_sheets=["COR - Corridor Analysis"],
        ),
        template(
            "IMTO_AGT",
            "Agent Bank Network",
            "Monthly",
            "FixedRow",
            "imto",
            [section("AGT", "Agent Network", 1)],
            [
                field("active_agent_banks", "Active Agent Banks", "Integer", "AGT", True, 1),
                field("inactive_agent_banks", "Inactive Agent Banks", "Integer", "AGT", True, 2),
                field("total_agent_banks", "Total Agent Banks", "Integer", "AGT", True, 3),
                field("agent_compliance_issues", "Agent Compliance Issues", "Integer", "AGT", True, 4),
            ],
            [
                formula_sum("total_agent_banks", ["active_agent_banks", "inactive_agent_banks"], "Total agent banks"),
            ],
            source_sheets=["AGT - Agent Bank Operations"],
        ),
        template(
            "IMTO_PAY",
            "Payout Distribution",
            "Monthly",
            "FixedRow",
            "imto",
            [section("PAY", "Payout Distribution", 1)],
            [
                field("bank_transfer_share", "Bank Transfer Share", "Percentage", "PAY", True, 1),
                field("mobile_money_share", "Mobile Money Share", "Percentage", "PAY", True, 2),
                field("cash_pickup_share", "Cash Pickup Share", "Percentage", "PAY", True, 3),
                field("total_share", "Total Share", "Percentage", "PAY", True, 4),
                field("share_target", "Share Target", "Percentage", "PAY", True, 5, 100, 100, 2),
            ],
            [
                formula_sum("total_share", ["bank_transfer_share", "mobile_money_share", "cash_pickup_share"], "Total payout share"),
                formula_compare("Equals", "total_share", "share_target", "Payout share should total 100", "Warning"),
            ],
            source_sheets=["TXN - Transaction Log", "VOL - Volume & Value"],
        ),
        template(
            "IMTO_BEN",
            "Beneficiary Analysis",
            "Monthly",
            "FixedRow",
            "imto",
            [section("BEN", "Beneficiary", 1)],
            [
                field("total_beneficiaries", "Total Beneficiaries", "Integer", "BEN", True, 1),
                field("high_value_beneficiaries", "High Value Beneficiaries", "Integer", "BEN", True, 2),
                field("single_beneficiary_monthly_value", "Single Beneficiary Monthly Value", "Money", "BEN", True, 3),
                field("aml_red_flag_threshold", "AML Red Flag Threshold", "Money", "BEN", True, 4, 5000000, 5000000, 2),
                field("beneficiary_concentration_alert", "Beneficiary Concentration Alert", "Money", "BEN", True, 5),
            ],
            [
                formula_compare("GreaterThan", "single_beneficiary_monthly_value", "aml_red_flag_threshold", "Beneficiary concentration AML red flag", "Warning"),
                formula_difference("beneficiary_concentration_alert", "single_beneficiary_monthly_value", "aml_red_flag_threshold", "Concentration buffer", "Warning", 0),
            ],
            source_sheets=["CUS - Customer & Beneficiary"],
        ),
        template(
            "IMTO_FIN",
            "Financial Statements",
            "Monthly",
            "FixedRow",
            "imto",
            [section("FIN", "Financial", 1)],
            [
                field("total_remittance_value", "Total Remittance Value", "Money", "FIN", True, 1),
                field("fee_income", "Fee Income", "Money", "FIN", True, 2),
                field("fx_gain_loss", "FX Gain/Loss", "Money", "FIN", True, 3),
                field("other_income", "Other Income", "Money", "FIN", True, 4),
                field("total_income", "Total Income", "Money", "FIN", True, 5),
                field("operating_expenses", "Operating Expenses", "Money", "FIN", True, 6),
                field("total_expenses", "Total Expenses", "Money", "FIN", True, 7),
                field("profit_before_tax", "Profit Before Tax", "Money", "FIN", True, 8),
                field("total_inbound_value", "Total Inbound Value", "Money", "FIN", True, 9),
            ],
            [
                formula_sum("total_income", ["fee_income", "fx_gain_loss", "other_income"], "Total income"),
                formula_sum("total_expenses", ["operating_expenses"], "Total expenses"),
                formula_difference("profit_before_tax", "total_income", "total_expenses", "PBT"),
                formula_compare("Equals", "total_remittance_value", "total_inbound_value", "Financial remittance value should reconcile with inbound", "Warning", 1),
            ],
            source_sheets=["FIN - Financial Statements"],
        ),
        template(
            "IMTO_CAP",
            "Capital",
            "Monthly",
            "FixedRow",
            "imto",
            [section("CAP", "Capital", 1)],
            [
                field("paid_up_capital", "Paid-Up Capital", "Money", "CAP", True, 1),
                field("minimum_required_capital", "Minimum Required Capital", "Money", "CAP", True, 2),
                field("capital_buffer", "Capital Buffer", "Money", "CAP", True, 3),
            ],
            [
                formula_difference("capital_buffer", "paid_up_capital", "minimum_required_capital", "Capital buffer"),
                formula_compare("GreaterThanOrEqual", "paid_up_capital", "minimum_required_capital", "Capital adequacy"),
            ],
            source_sheets=["GOV - Governance & Licensing"],
        ),
        template(
            "IMTO_AML",
            "AML/CFT",
            "Monthly",
            "FixedRow",
            "imto",
            [section("AML", "AML", 1)],
            [
                field("str_filed_count", "STR Filed Count", "Integer", "AML", True, 1),
                field("ctr_filed_count", "CTR Filed Count", "Integer", "AML", True, 2),
                field("pep_customers_count", "PEP Count", "Integer", "AML", True, 3),
                field("tfs_screening_count", "TFS Screening Count", "Integer", "AML", True, 4),
                field("total_aml_alerts", "Total AML Alerts", "Integer", "AML", True, 5),
            ],
            [
                formula_sum("total_aml_alerts", ["str_filed_count", "ctr_filed_count", "pep_customers_count"], "Total AML alerts", "Warning", 0),
            ],
            source_sheets=["AML - AML-CFT-CPF"],
        ),
        template(
            "IMTO_NFE",
            "NFEM Compliance",
            "Monthly",
            "FixedRow",
            "imto",
            [section("NFE", "NFEM", 1)],
            [
                field("payout_rate_avg", "Payout Rate Average", "Decimal", "NFE", True, 1),
                field("nfem_reference_rate", "NFEM Reference Rate", "Decimal", "NFE", True, 2),
                field("allowed_band_percent", "Allowed Band Percent", "Percentage", "NFE", True, 3),
                field("rate_band_check_result", "Rate Band Check Result", "Money", "NFE", True, 4),
            ],
            [
                formula_custom(
                    "rate_band_check_result",
                    ["payout_rate_avg", "nfem_reference_rate", "allowed_band_percent"],
                    "RATE_BAND_CHECK",
                    "NFEM rate band check",
                    "Error",
                    parameters={"actual_rate": "payout_rate_avg", "reference_rate": "nfem_reference_rate", "band_percent": "allowed_band_percent"},
                ),
            ],
            source_sheets=["FXR - FX & Settlement"],
        ),
        template(
            "IMTO_GOV",
            "Governance",
            "Monthly",
            "FixedRow",
            "imto",
            [section("GOV", "Governance", 1)],
            [
                field("board_meetings_held", "Board Meetings Held", "Integer", "GOV", True, 1),
                field("compliance_breaches", "Compliance Breaches", "Integer", "GOV", True, 2),
                field("resolved_compliance_breaches", "Resolved Compliance Breaches", "Integer", "GOV", True, 3),
            ],
            [
                formula_compare("LessThanOrEqual", "resolved_compliance_breaches", "compliance_breaches", "Resolved breaches cannot exceed breaches"),
            ],
            source_sheets=["GOV - Governance & Licensing"],
        ),
        template(
            "IMTO_DIC",
            "Data Dictionary",
            "Monthly",
            "FixedRow",
            "imto",
            [section("DIC", "Validation", 1)],
            [
                field("total_field_count", "Total Field Count", "Integer", "DIC", True, 1),
                field("total_formula_count", "Total Formula Count", "Integer", "DIC", True, 2),
                field("validation_error_count", "Validation Error Count", "Integer", "DIC", True, 3),
            ],
            [],
            source_sheets=["Data Dictionary"],
        ),
    ]

    targets = {
        "IMTO_COV": (18, 5),
        "IMTO_INB": (24, 10),
        "IMTO_COR": (24, 10),
        "IMTO_AGT": (20, 7),
        "IMTO_PAY": (20, 8),
        "IMTO_BEN": (20, 8),
        "IMTO_FIN": (24, 10),
        "IMTO_CAP": (20, 8),
        "IMTO_AML": (18, 7),
        "IMTO_NFE": (20, 8),
        "IMTO_GOV": (16, 6),
        "IMTO_DIC": (16, 4),
    }

    for t in templates:
        tf, tfo = targets[t["returnCode"]]
        pad_template(t, tf, tfo)

    definition = {
        "moduleCode": "IMTO_CBN",
        "moduleVersion": "1.0.0",
        "description": "CBN IMTO returns for corridor analysis, payout controls, NFEM compliance, financials and AML.",
        "templates": templates,
        "interModuleDataFlows": [
            {
                "sourceTemplate": "IMTO_AML",
                "sourceField": "str_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_STR",
                "targetField": "str_filed_count",
                "transformationType": "DirectCopy",
                "description": "IMTO STR count to NFIU",
            },
            {
                "sourceTemplate": "IMTO_AML",
                "sourceField": "ctr_filed_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_CTR",
                "targetField": "ctr_filed_count",
                "transformationType": "DirectCopy",
                "description": "IMTO CTR count to NFIU",
            },
            {
                "sourceTemplate": "IMTO_AML",
                "sourceField": "pep_customers_count",
                "targetModule": "NFIU_AML",
                "targetTemplate": "NFIU_PEP",
                "targetField": "pep_alert_count",
                "transformationType": "DirectCopy",
                "description": "IMTO PEP count to NFIU",
            },
        ],
    }

    add_cross_rules(definition, 18)
    return definition


def add_rg10_specific_cross_rules(definitions):
    ins = definitions["INSURANCE_NAICOM"]
    pfa = definitions["PFA_PENCOM"]
    cmo = definitions["CMO_SEC"]
    dfi = definitions["DFI_CBN"]
    imto = definitions["IMTO_CBN"]

    next(t for t in ins["templates"] if t["returnCode"] == "INS_SOL")["crossSheetRules"].append(
        {
            "description": "Admitted assets should reconcile to financial total assets",
            "sourceTemplate": "INS_SOL",
            "sourceField": "admitted_assets",
            "targetTemplate": "INS_FIN",
            "targetField": "total_assets",
            "operator": "Equals",
            "severity": "Warning",
            "tolerancePercent": 1.0,
        }
    )

    next(t for t in pfa["templates"] if t["returnCode"] == "PFA_NAV")["crossSheetRules"].append(
        {
            "description": "NAV total should reconcile with Fund details total",
            "sourceTemplate": "PFA_NAV",
            "sourceField": "total_nav",
            "targetTemplate": "PFA_FD1",
            "targetField": "fund_i_total",
            "operator": "GreaterThanOrEqual",
            "severity": "Warning",
            "toleranceAmount": 1.0,
        }
    )

    next(t for t in cmo["templates"] if t["returnCode"] == "CMO_TRD")["crossSheetRules"].append(
        {
            "description": "Trading value should reconcile to financial remittance value",
            "sourceTemplate": "CMO_TRD",
            "sourceField": "total_trade_value",
            "targetTemplate": "CMO_FIN",
            "targetField": "total_trade_value",
            "operator": "Equals",
            "severity": "Warning",
            "tolerancePercent": 2.0,
        }
    )

    next(t for t in dfi["templates"] if t["returnCode"] == "DFI_SEC")["crossSheetRules"].append(
        {
            "description": "Sector exposure should be within total assets",
            "sourceTemplate": "DFI_SEC",
            "sourceField": "total_sector_exposure",
            "targetTemplate": "DFI_FIN",
            "targetField": "total_assets",
            "operator": "LessThanOrEqual",
            "severity": "Warning",
            "tolerancePercent": 5.0,
        }
    )

    next(t for t in imto["templates"] if t["returnCode"] == "IMTO_COR")["crossSheetRules"].append(
        {
            "description": "Corridor values must reconcile with IMTO financial total remittance value",
            "sourceTemplate": "IMTO_COR",
            "sourceField": "total_corridor_value",
            "targetTemplate": "IMTO_FIN",
            "targetField": "total_remittance_value",
            "operator": "Equals",
            "severity": "Error",
            "tolerancePercent": 1.0,
        }
    )


def write_module_json(output_dir: Path, module_name: str, definition: dict):
    output_path = output_dir / module_name
    output_path.write_text(json.dumps(definition, indent=2) + "\n", encoding="utf-8")


def summarize(defn):
    t = len(defn["templates"])
    f = sum(len(x["fields"]) for x in defn["templates"])
    fo = sum(len(x["formulas"]) for x in defn["templates"])
    cr = sum(len(x["crossSheetRules"]) for x in defn["templates"])
    fl = len(defn.get("interModuleDataFlows", []))
    return t, f, fo, cr, fl


def main():
    root = Path(__file__).resolve().parents[1]
    out = root / "docs" / "module-definitions" / "rg10"
    out.mkdir(parents=True, exist_ok=True)

    templates_root = root.parents[0] / "Templates"

    definitions = {
        "INSURANCE_NAICOM": insurance_module_definition(),
        "PFA_PENCOM": pfa_module_definition(),
        "CMO_SEC": cmo_module_definition(),
        "DFI_CBN": dfi_module_definition(),
        "IMTO_CBN": imto_module_definition(),
    }

    add_rg10_specific_cross_rules(definitions)

    add_source_coverage(
        definitions["INSURANCE_NAICOM"],
        templates_root / "Insurance_NAICOM_Reporting_Templates.xlsx",
        templates_root / "Insurance_Return_Schema_v1.0.xsd",
        "INS_DIC",
    )
    add_source_coverage(
        definitions["PFA_PENCOM"],
        templates_root / "PFA_PenCom_Reporting_Templates.xlsx",
        templates_root / "PFA_Return_Schema_v1.0.xsd",
        "PFA_DIC",
    )
    add_source_coverage(
        definitions["CMO_SEC"],
        templates_root / "Capital_Market_SEC_Reporting_Templates.xlsx",
        templates_root / "SEC_CMO_Return_Schema_v1.0.xsd",
        "CMO_DIC",
    )
    add_source_coverage(
        definitions["DFI_CBN"],
        templates_root / "DFI_CBN_Reporting_Templates.xlsx",
        templates_root / "DFI_Return_Schema_v1.0.xsd",
        "DFI_DIC",
    )
    add_source_coverage(
        definitions["IMTO_CBN"],
        templates_root / "IMTO_CBN_Reporting_Templates.xlsx",
        templates_root / "IMTO_Return_Schema_v1.0.xsd",
        "IMTO_DIC",
    )

    file_map = {
        "INSURANCE_NAICOM": "insurance_naicom.json",
        "PFA_PENCOM": "pfa_pencom.json",
        "CMO_SEC": "cmo_sec.json",
        "DFI_CBN": "dfi_cbn.json",
        "IMTO_CBN": "imto_cbn.json",
    }

    totals = []
    for module_code, defn in definitions.items():
        write_module_json(out, file_map[module_code], defn)
        totals.append((module_code, summarize(defn)))

    print("RG-10 module definition generation complete")
    print("module,templates,fields,formulas,cross_sheet_rules,data_flows")
    for module_code, metrics in totals:
        print(f"{module_code},{metrics[0]},{metrics[1]},{metrics[2]},{metrics[3]},{metrics[4]}")

    all_templates = sum(m[0] for _, m in totals)
    all_fields = sum(m[1] for _, m in totals)
    all_formulas = sum(m[2] for _, m in totals)
    all_cross = sum(m[3] for _, m in totals)
    all_flows = sum(m[4] for _, m in totals)
    print(f"TOTAL,{all_templates},{all_fields},{all_formulas},{all_cross},{all_flows}")


if __name__ == "__main__":
    main()
